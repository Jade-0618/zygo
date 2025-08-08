# app/langgraph_def/graph_builder.py
# -*- coding: utf-8 -*-

# =================================================================================
# 1. Imports & Setup (无变动)
# =================================================================================
import json
import sys
import os
import re
import shutil
import subprocess
from copy import deepcopy
from pathlib import Path
from pprint import pprint
import time
import textwrap
import socket
from typing import Dict, Optional, List

from langchain_core.messages import HumanMessage
from langchain_openai import ChatOpenAI
from langgraph.graph import StateGraph, END
from langchain_core.output_parsers import JsonOutputParser
from langchain_core.prompts import ChatPromptTemplate, SystemMessagePromptTemplate, HumanMessagePromptTemplate
from thefuzz import process
from pydantic import SecretStr

from .agent_state import AgentState
from app.models import User, Device
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# =================================================================================
# 2. 模型初始化 (修改为使用环境变量)
# =================================================================================
API_KEY_STR = os.environ.get('LANGCHAIN_API_KEY')
BASE_URL = os.environ.get('LANGCHAIN_BASE_URL', "https://ark.cn-beijing.volces.com/api/v3")

# 检查API_KEY是否设置
if not API_KEY_STR:
    raise ValueError("LANGCHAIN_API_KEY environment variable is not set. Please set it before running the application.")

# 转换为SecretStr类型
API_KEY = SecretStr(API_KEY_STR)

system_architect_model = ChatOpenAI(model="ep-20250223112748-lthgv", temperature=0.5, api_key=API_KEY,
                                    base_url=BASE_URL)
module_architect_model = ChatOpenAI(model="ep-20250223112748-lthgv", temperature=0.4, api_key=API_KEY,
                                    base_url=BASE_URL)
api_designer_model = ChatOpenAI(model="ep-20250223112748-lthgv", temperature=0.3, api_key=API_KEY, base_url=BASE_URL)
developer_model = ChatOpenAI(model="ep-20250223112748-lthgv", temperature=0.2, api_key=API_KEY, base_url=BASE_URL)
tester_model = ChatOpenAI(model="ep-20250223112748-lthgv", temperature=0.1, api_key=API_KEY, base_url=BASE_URL)
dp_extractor_model = ChatOpenAI(model="ep-20250223112748-lthgv", temperature=0.2, api_key=API_KEY, base_url=BASE_URL)


# =================================================================================
# 3. 辅助函数及核心模块生成器 (无变动)
# =================================================================================

def robust_rmtree(path: Path, retries: int = 5, delay: int = 1):
    """
    一个更健壮的删除函数，可以处理文件和目录，并在遇到权限错误时重试。
    """
    for i in range(retries):
        if not path.exists():
            return

        try:
            if path.is_dir():
                shutil.rmtree(path)
            else:
                path.unlink()

            if not path.exists():
                return

        except (PermissionError, OSError) as e:
            print(
                f"  [robust_rmtree] Warn: Attempt {i + 1}/{retries} to delete '{path}' failed: {e}. Retrying in {delay}s...")
            time.sleep(delay)

    if path.exists():
        raise OSError(f"Failed to delete path '{path}' after {retries} retries. It might be locked by another process.")


API_PACKAGE_DIR = Path(__file__).resolve().parent.parent.parent / 'API_Package'

BOARD_ID_MAP = {
    "esp32": "esp32dev",
    "esp32dev": "esp32dev",
    "esp32-wroom-32": "esp32dev",
    "esp32-wroom-32d": "esp32dev",
    "esp32-wroom-32u": "esp32dev",
    "esp32s3": "esp32-s3-devkitc-1",
    "esp32s2": "esp32-s2-saola-1",
    "esp32c3": "esp32-c3-devkitm-1",
}


def find_board_id(user_board_name: str) -> Optional[str]:
    """使用映射表和模糊搜索来查找最匹配的官方开发板ID。"""
    if not user_board_name: return None
    clean_name = user_board_name.strip().lower().replace("-", "").replace("_", "")

    for alias, official_id in BOARD_ID_MAP.items():
        if clean_name == alias.replace("-", "").replace("_", ""):
            print(f"  通过精确匹配找到: '{user_board_name}' -> '{official_id}'")
            return official_id

    best_match, score = process.extractOne(clean_name, BOARD_ID_MAP.keys())
    if score > 85:
        matched_id = BOARD_ID_MAP[best_match]
        print(f"  通过模糊搜索找到匹配: '{user_board_name}' -> '{best_match}' (相似度: {score}%) -> '{matched_id}'")
        return matched_id

    print(f"  在本地映射中未找到 '{user_board_name}' 的高可信度匹配。将使用默认值。")
    return None


def find_api_spec(peripheral_name: str) -> Optional[str]:
    """
    通过两阶段流程在本地查找外设的API规范：
    1. 精确文件名匹配。
    2. 模糊文件名匹配。
    如果找到，返回文件内容；否则返回 None。
    """
    if not API_PACKAGE_DIR.exists():
        print(f"  警告: API包目录 '{API_PACKAGE_DIR}' 不存在。跳过本地搜索。")
        return None

    exact_filename = f"{peripheral_name.upper().replace(' ', '_')}_API_Package.json"
    exact_filepath = API_PACKAGE_DIR / exact_filename
    if exact_filepath.is_file():
        print(f"  API found via exact match in local cache: '{exact_filename}'")
        try:
            return exact_filepath.read_text(encoding='utf-8')
        except Exception as e:
            print(f"  错误: 读取缓存的API文件失败 {exact_filepath}: {e}")
            return None

    all_api_files = [f.name for f in API_PACKAGE_DIR.glob('*_API_Package.json')]
    if not all_api_files:
        return None

    file_map = {f.replace('_API_Package.json', ''): f for f in all_api_files}
    best_match_core, score = process.extractOne(peripheral_name.upper(), file_map.keys())

    CONFIDENCE_THRESHOLD = 85
    if score >= CONFIDENCE_THRESHOLD:
        fuzzy_filename = file_map[best_match_core]
        fuzzy_filepath = API_PACKAGE_DIR / fuzzy_filename
        print(f"  API found via fuzzy match (Confidence: {score}%): '{peripheral_name}' -> '{fuzzy_filename}'")
        try:
            return fuzzy_filepath.read_text(encoding='utf-8')
        except Exception as e:
            print(f"  错误: 读取缓存的API文件失败 {fuzzy_filepath}: {e}")
            return None

    return None


def get_local_ip() -> str:
    s = None
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(('8.8.8.8', 53))
        ip = s.getsockname()[0]
        return ip
    except Exception as e:
        return "127.0.0.1"
    finally:
        if s: s.close()


def extract_code(content: str, lang: str = "cpp", block_name: str = None) -> str:
    if block_name:
        pattern = re.compile(rf'\[{block_name.upper()}\]\s*```{lang}\s*([\s\S]*?)\s*```', re.DOTALL)
    else:
        pattern = re.compile(rf'```{lang}\s*([\s\S]*?)\s*```', re.DOTALL)
    match = pattern.search(content)
    return match.group(1).strip() if match else f"// Error: Could not extract code block '{block_name or lang}'."


def generate_mbedtls_sha256_header() -> str:
    return """
#ifndef CUSTOM_SHA256_H
#define CUSTOM_SHA256_H
#include <Arduino.h>
#include "mbedtls/sha256.h"
class SHA256 {
public:
    SHA256();
    void update(const void *data, size_t len);
    void finalize(byte *hash);
    static String toString(const byte* hash, int len = 32);
private:
    mbedtls_sha256_context ctx;
};
#endif // CUSTOM_SHA256_H
"""


def generate_mbedtls_sha256_source() -> str:
    return """
#include "SHA256.h"
SHA256::SHA256() {
    mbedtls_sha256_init(&ctx);
    mbedtls_sha256_starts_ret(&ctx, 0);
}
void SHA256::update(const void *data, size_t len) {
    if (len == 0) return;
    mbedtls_sha256_update_ret(&ctx, (const unsigned char *)data, len);
}
void SHA256::finalize(byte *hash) {
    mbedtls_sha256_finish_ret(&ctx, hash);
    mbedtls_sha256_starts_ret(&ctx, 0);
}
String SHA256::toString(const byte* hash, int len) {
    char hex_string[len * 2 + 1];
    for (int i = 0; i < len; i++) {
        sprintf(&hex_string[i * 2], "%02x", hash[i]);
    }
    hex_string[len * 2] = '\\0';
    return String(hex_string);
}
"""


def generate_tuya_handler_header() -> str:
    return """
#ifndef TUYA_HANDLER_H
#define TUYA_HANDLER_H
#include <Arduino.h>
#include <WiFiClientSecure.h>
#include <PubSubClient.h>

// Define a function pointer type for the application callback
typedef void (*TuyaAppCallback)(String &topic, String &payload);

// Provides the calculated credentials for app_main to use in its connection logic
void tuya_get_mqtt_credentials(char* out_client_id, char* out_username, char* out_password);

// A non-blocking, passive initialization function
void tuya_init(WiFiClientSecure& wifiClient, PubSubClient& mqttClient, TuyaAppCallback app_callback);

// A function to be called by app_main AFTER a successful connection
void tuya_subscribe_topics();

void tuya_loop();

// Publishes data to the Tuya cloud
bool tuya_publish_data(const String& data_json_string);

// A function for app_main's master callback to dispatch Tuya messages
void tuya_handle_mqtt_message(char *topic, byte *payload, unsigned int length);

#endif // TUYA_HANDLER_H
"""


def generate_config_manager_header(
    device_id: str,
    wifi_ssid: str,
    wifi_password: str,
    product_id: str,
    cloud_device_id: str,
    device_secret: str
) -> str:
    local_pc_ip = get_local_ip()

    # 使用传入的参数，如果为空则使用占位符
    final_wifi_ssid = wifi_ssid if wifi_ssid else "YOUR_WIFI_SSID"
    final_wifi_password = wifi_password if wifi_password else "YOUR_WIFI_PASSWORD"
    final_product_id = product_id if product_id else "YOUR_PRODUCT_ID"
    final_cloud_device_id = cloud_device_id if cloud_device_id else "YOUR_DEVICE_ID"
    final_device_secret = device_secret if device_secret else "YOUR_DEVICE_SECRET"

    print("--- [generate_config_manager_header] Generating with following data: ---")
    print(f"  WIFI_SSID: {final_wifi_ssid}")
    print(f"  WIFI_PASSWORD: {'*' * len(final_wifi_password) if final_wifi_password else '(empty)'}")
    print(f"  TUYA_PRODUCT_ID: {final_product_id}")
    print(f"  TUYA_DEVICE_ID: {final_cloud_device_id}")
    print(f"  TUYA_DEVICE_SECRET: {'*' * len(final_device_secret) if final_device_secret else '(empty)'}")
    print("---------------------------------------------------------------------")

    return f'''
#ifndef CONFIG_MANAGER_H
#define CONFIG_MANAGER_H

// --- Wi-Fi Credentials ---
#define WIFI_SSID "{final_wifi_ssid}"
#define WIFI_PASSWORD "{final_wifi_password}"

// --- Tuya Cloud Credentials ---
#define TUYA_PRODUCT_ID "{final_product_id}"
#define TUYA_DEVICE_ID  "{final_cloud_device_id}"
#define TUYA_DEVICE_SECRET "{final_device_secret}"

// --- Local Network Configuration ---
#define MQTT_BROKER "{local_pc_ip}"
#define MQTT_PORT 1883
#define OTA_HTTP_SERVER "{local_pc_ip}"
#define OTA_HTTP_PORT 8000

// --- Device Identity ---
#define DEVICE_ID "{device_id}"
#define FIRMWARE_VERSION "1.0.0"

// --- MQTT Topics ---
#define OTA_TOPIC_BASE "/ota/"
#define DEBUG_TOPIC_BASE "/debug/"

#endif // CONFIG_MANAGER_H
'''


def generate_tuya_handler_source() -> str:
    # V2 Refactored: This handler is now a passive library.
    # It provides credentials but does not manage the connection itself.
    # FIX: Corrected the formatting of the multi-line CA certificate string to resolve C++ compilation errors.
    return """
#include "tuya_handler.h"
#include "config_manager.h"
#include <ArduinoJson.h>
#include <time.h>
#include "SHA256.h"

// The root CA certificate for Tuya's MQTT broker
static const char tuya_ca_cert[] PROGMEM =
    "-----BEGIN CERTIFICATE-----\\n"
    "MIIGiTCCBXGgAwIBAgIIGkbwkRaiCBgwDQYJKoZIhvcNAQELBQAwgbQxCzAJBgNV\\n"
    "BAYTAlVTMRAwDgYDVQQIEwdBcml6b25hMRMwEQYDVQQHEwpTY290dHNkYWxlMRow\\n"
    "GAYDVQQKExFHb0RhZGR5LmNvbSwgSW5jLjEtMCsGA1UECxMkaHR0cDovL2NlcnRz\\n"
    "LmdvZGFkZHkuY29tL3JlcG9zaXRvcnkvMTMwMQYDVQQDEypHbyBEYWRkeSBTZWN1\\n"
    "cmUgQ2VydGlmaWNhdGUgQXV0aG9yaXR5IC0gRzIwHhcNMjQwODE5MDU1NTM4WhcN\\n"
    "MjUwOTIwMDU1NTM4WjAXMRUwEwYDVQQDDAwqLnR1eWFjbi5jb20wggEiMA0GCSqG\\n"
    "SIb3DQEBAQUAA4IBDwAwggEKAoIBAQDlYq+PYMUih5G0Ob9XX1a57li+CA2YCy38\\n"
    "1gmpStB+/XqC4mHc8GYhEV9rnfd0egs8R6G9J/FwXw0UfNER3UEg1WYEJ0Hi6eMX\\n"
    "0BI65+wZdvJxEpFhwcXU50tPTADxudw8I5haJ5Cv453yH7/kg2M2Qk32YjLwV9Yz\\n"
    "79c6Ogzsg27FCDTghiWuqMq3cImNcYGKC0vNv5D8B+YjI41n1a0hgZXloP9478b/\\n"
    "S/uxPZdg8CpsXRpcTwmxcOScy7ip4aqiYjvjwVB2ZIhprTJiwdInWfiUYitgD5j+\\n"
    "JzW1hahMwu3fgYYozcduYdbbTyicAbGN88EvT1XChTEun72Tw0JfAgMBAAGjggM5\\n"
    "MIIDNTAMBgNVHRMBAf8EAjAAMB0GA1UdJQQWMBQGCCsGAQUFBwMBBggrBgEFBQcD\\n"
    "AjAOBgNVHQ8BAf8EBAMCBaAwOQYDVR0fBDIwMDAuoCygKoYoaHR0cDovL2NybC5n\\n"
    "b2RhZGR5LmNvbS9nZGlnMnMxLTI4ODM0LmNybDBdBgNVHSAEVjBUMEgGC2CGSAGG\\n"
    "/W0BBxcBMDkwNwYIKwYBBQUHAgEWK2h0dHA6Ly9jZXJ0aWZpY2F0ZXMuZ29kYWRk\\n"
    "eS5jb20vcmVwb3NpdG9yeS8wCAYGZ4EMAQIBMHYGCCsGAQUFBwEBBGowaDAkBggr\\n"
    "BgEFBQcwAYYYaHR0cDovL29jc3AuZ29kYWRkeS5jb20vMEAGCCsGAQUFBzAChjRo\\n"
    "dHRwOi8vY2VydGlmaWNhdGVzLmdvZGFkZHkuY29tL3JlcG9zaXRvcnkvZ2RpZzIu\\n"
    "Y3J0MB8GA1UdIwQYMBaAFEDCvSeOzDSDMKIz1/tss/C0LIDOMCMGA1UdEQQcMBqC\\n"
    "DCoudHV5YWNuLmNvbYIKdHV5YWNuLmNvbTAdBgNVHQ4EFgQUGwrpXqEzmkB903gf\\n"

    "iQpZ18e8geMwggF9BgorBgEEAdZ5AgQCBIIBbQSCAWkBZwB2ABLxTjS9U3JMhAYZ\\n"
    "w48/ehP457Vih4icbTAFhOvlhiY6AAABkWk0kgkAAAQDAEcwRQIgfAhWc0NWPQFk\\n"
    "KdCfg4A9A+Io0bWQdAKr6/vYpr3IQaYCIQCbYHiULR0Nkw4cGtFK3HympmuNbgkt\\n"
    "rd51XUQcTfCwTAB1AH1ZHhLheCp7HGFnfF79+NCHXBSgTpWeuQMv2Q6MLnm4AAAB\\n"
    "kWk0ktwAAAQDAEYwRAIgKM9rEIVMjHCUnxUkQYgXeVvVume85E6oiHoFfBaGuIEC\\n"
    "IER4giiSxqR4ftNJkfi8v4ftQrrOt7iZ4FDlnSzKBCpLAHYAzPsPaoVxCWX+lZtT\\n"
    "zumyfCLphVwNl422qX5UwP5MDbAAAAGRaTSTmwAABAMARzBFAiBCqgaBigm1c/hH\\n"
    "owy25qZfn+I8mpc+H1VrVlEZZqAj8gIhAI2ZV/CaarGQ/j8HRwHq7vO+5j/QlXbJ\\n"
    "tI/XHIHqZQW2MA0GCSqGSIb3DQEBCwUAA4IBAQAnWV1if9nZK6aVftzj/w2VkmBY\\n"
    "zBLSO+3Co1Qyc3qxBsCpdLxVCycN9HcmOAAgVdMg5WLs542KGMvIahh0PJzyIrMG\\n"
    "uTeLUvUmb9yGZb+oDLlsqLeAxJZi/Mf4ZN5Ezq52bDotXb6+qrftCrQj+Vz3dp9N\\n"
    "U9XGvts/lM1dpnnoCoVpMTM+kzyzkmIJbb/zSy8U1TLbja5HYdtYVodeMexG+PE/\\n"
    "F+OGeB3AWU5yhSr65XRMWKynNglfspsnvU2azab+3CViOFsCR6Th30ohQKgxjldQ\\n"
    "xfN2SPdPvZjOnmXZT75rMeGahN8PqloYFP12VwsF+IPo3m50U2hstS2IiA3U\\n"
    "-----END CERTIFICATE-----\\n"
    "-----BEGIN CERTIFICATE-----\\n"
    "MIIE0DCCA7igAwIBAgIBBzANBgkqhkiG9w0BAQsFADCBgzELMAkGA1UEBhMCVVMx\\n"
    "EDAOBgNVBAgTB0FyaXpvbmExEzARBgNVBAcTClNjb3R0c2RhbGUxGjAYBgNVBAoT\\n"
    "EUdvRGFkZHkuY29tLCBJbmMuMTEwLwYDVQQDEyhHbyBEYWRkeSBSb290IENlcnRp\\n"
    "ZmljYXRlIEF1dGhvcml0eSAtIEcyMB4XDTExMDUwMzA3MDAwMFoXDTMxMDUwMzA3\\n"
    "MDAwMFowgbQxCzAJBgNVBAYTAlVTMRAwDgYDVQQIEwdBcml6b25hMRMwEQYDVQQH\\n"
    "EwpTY290dHNkYWxlMRowGAYDVQQKExFHb0RhZGR5LmNvbSwgSW5jLjEtMCsGA1UE\\n"
    "CxMkaHR0cDovL2NlcnRzLmdvZGFkZHkuY29tL3JlcG9zaXRvcnkvMTMwMQYDVQQD\\n"
    "EypHbyBEYWRkeSBTZWN1cmUgQ2VydGlmaWNhdGUgQXV0aG9yaXR5IC0gRzIwggEi\\n"
    "MA0GCSqGSIb3DQEBAQUAA4IBDwAwggEKAoIBAQC54MsQ1K92vdSTYuswZLiBCGzD\\n"
    "BNliF44v/z5lz4/OYuY8UhzaFkVLVat4a2ODYpDOD2lsmcgaFItMzEUz6ojcnqOv\\n"
    "K/6AYZ15V8TPLvQ/MDxdR/yaFrzDN5ZBUY4RS1T4KL7QjL7wMDge87Am+GZHY23e\\n"
    "cSZHjzhHU9FGHbTj3ADqRay9vHHZqm8A29vNMDp5T19MR/gd71vCxJ1gO7GyQ5HY\\n"
    "pDNO6rPWJ0+tJYqlxvTV0KaudAVkV4i1RFXULSo6Pvi4vekyCgKUZMQWOlDxSq7n\\n"
    "eTOvDCAHf+jfBDnCaQJsY1L6d8EbyHSHyLmTGFBUNUtpTrw700kuH9zB0lL7AgMB\\n"
    "AAGjggEaMIIBFjAPBgNVHRMBAf8EBTADAQH/MA4GA1UdDwEB/wQEAwIBBjAdBgNV\\n"
    "HQ4EFgQUQMK9J47MNIMwojPX+2yz8LQsgM4wHwYDVR0jBBgwFoAUOpqFBxBnKLbv\\n"
    "9r0FQW4gwZTaD94wNAYIKwYBBQUHAQEEKDAmMCQGCCsGAQUFBzABhhhodHRwOi8v\\n"
    "b2NzcC5nb2RhZGR5LmNvbS8wNQYDVR0fBC4wLDAqoCigJoYkaHR0cDovL2NybC5n\\n"
    "b2RhZGR5LmNvbS9nZHJvb3QtZzIuY3JsMEYGA1UdIAQ/MD0wOwYEVR0gADAzMDEG\\n"

    "CCsGAQUFBwIBFiVodHRwczovL2NlcnRzLmdvZGFkZHkuY29tL3JlcG9zaXRvcnkv\\n"
    "MA0GCSqGSIb3DQEBCwUAA4IBAQAIfmyTEMg4uJapkEv/oV9PBO9sPpyIBslQj6Zz\\n"
    "91cxG7685C/b+LrTW+C05+Z5Yg4MotdqY3MxtfWoSKQ7CC2iXZDXtHwlTxFWMMS2\\n"
    "RJ17LJ3lXubvDGGqv+QqG+6EnriDfcFDzkSnE3ANkR/0yBOtg2DZ2HKocyQetawi\\n"
    "DsoXiWJYRBuriSUBAA/NxBti21G00w9RKpv0vHP8ds42pM3Z2Czqrpv1KrKQ0U11\\n"
    "GIo/ikGQI31bS/6kA1ibRrLDYGCD+H1QQc7CoZDDu+8CL9IVVO5EFdkKrqeKM+2x\\n"
    "LXY2JtwE65/3YR8V3Idv7kaWKK2hJn0KCacuBKONvPi8BDAB\\n"
    "-----END CERTIFICATE-----\\n"
    "-----BEGIN CERTIFICATE-----\\n"
    "MIIEfTCCA2WgAwIBAgIDG+cVMA0GCSqGSIb3DQEBCwUAMGMxCzAJBgNVBAYTAlVT\\n"
    "MSEwHwYDVQQKExhUaGUgR28gRGFkZHkgR3JvdXAsIEluYy4xMTAvBgNVBAsTKEdv\\n"
    "IERhZGR5IENsYXNzIDIgQ2VydGlmaWNhdGlvbiBBdXRob3JpdHkwHhcNMTQwMTAx\\n"
    "MDcwMDAwWhcNMzEwNTMwMDcwMDAwWjCBgzELMAkGA1UEBhMCVVMxEDAOBgNVBAgT\\n"
    "B0FyaXpvbmExEzARBgNVBAcTClNjb3R0c2RhbGUxGjAYBgNVBAoTEUdvRGFkZHku\\n"
    "Y29tLCBJbmMuMTEwLwYDVQQDEyhHbyBEYWRkeSBSb290IENlcnRpZmljYXRlIEF1\\n"
    "dGhvcml0eSAtIEcyMIIBIjANBgkqhkiG9w0BAQEFAAOCAQ8AMIIBCgKCAQEAv3Fi\\n"
    "CPH6WTT3G8kYo/eASVjpIoMTpsUgQwE7hPHmhUmfJ+r2hBtOoLTbcJjHMgGxBT4H\\n"
    "Tu70+k8vWTAi56sZVmvigAf88xZ1gDlRe+X5NbZ0TqmNghPktj+pA4P6or6KFWp/\\n"
    "3gvDthkUBcrqw6gElDtGfDIN8wBmIsiNaW02jBEYt9OyHGC0OPoCjM7T3UYH3go+\\n"
    "6118yHz7sCtTpJJiaVElBWEaRIGMLKlDliPfrDqBmg4pxRyp6V0etp6eMAo5zvGI\\n"
    "gPtLXcwy7IViQyU0AlYnAZG0O3AqP26x6JyIAX2f1PnbU21gnb8s51iruF9G/M7E\\n"
    "GwM8CetJMVxpRrPgRwIDAQABo4IBFzCCARMwDwYDVR0TAQH/BAUwAwEB/zAOBgNV\\n"
    "HQ8BAf8EBAMCAQYwHQYDVR0OBBYEFDqahQcQZyi27/a9BUFuIMGU2g/eMB8GA1Ud\\n"
    "IwQYMBaAFNLEsNKR1EwRcbNhyz2h/t2oatTjMDQGCCsGAQUFBwEBBCgwJjAkBggr\\n"
    "BgEFBQcwAYYYaHR0cDovL29jc3AuZ29kYWRkeS5jb20vMDIGA1UdHwQrMCkwJ6Al\\n"
    "oCOGIWh0dHA6Ly9jcmwuZ29kYWRkeS5jb20vZ2Ryb290LmNybDBGBgNVHSAEPzA9\\n"
    "MDsGBFUdIAAwMzAxBggrBgEFBQcCARYlaHR0cHM6Ly9jZXJ0cy5nb2RhZGR5LmNv\\n"
    "bS9yZXBvc2l0b3J5LzANBgkqhkiG9w0BAQsFAAOCAQEAWQtTvZKGEacke+1bMc8d\\n"
    "H2xwxbhuvk679r6XUOEwf7ooXGKUwuN+M/f7QnaF25UcjCJYdQkMiGVnOQoWCcWg\\n"
    "OJekxSOTP7QYpgEGRJHjp2kntFolfzq3Ms3dhP8qOCkzpN1nsoX+oYggHFCJyNwq\\n"
    "9kIDN0zmiN/VryTyscPfzLXs4Jlet0lUIDyUGAzHHFIYSaRt4bNYC8nY7NmuHDKO\\n"
    "KHAN4v6mF56ED71XcLNa6R+ghlO773z/aQvgSMO3kwvIClTErF0UZzdsyqUvMQg3\\n"
    "qm5vjLyb4lddJIGvl5echK1srDdMZvNhkREg5L4wn3qkKQmw4TRfZHcYQFHfjDCm\\n"
    "rw==\\n"
    "-----END CERTIFICATE-----\\n";

static PubSubClient* _mqttClient;
static TuyaAppCallback _app_callback = nullptr;

static String hmac256(const char* key, size_t key_len, const char* message, size_t msg_len) {
    SHA256 sha;
    byte k_ipad[64], k_opad[64];
    memset(k_ipad, 0, sizeof(k_ipad));
    memset(k_opad, 0, sizeof(k_opad));
    memcpy(k_ipad, key, key_len);
    memcpy(k_opad, key, key_len);
    for (int i = 0; i < 64; i++) {
        k_ipad[i] ^= 0x36;
        k_opad[i] ^= 0x5c;
    }
    sha.update(k_ipad, sizeof(k_ipad));
    sha.update(message, msg_len);
    byte hmac[32];
    sha.finalize(hmac);
    sha.update(k_opad, sizeof(k_opad));
    sha.update(hmac, sizeof(hmac));
    sha.finalize(hmac);
    return SHA256::toString(hmac);
}

// This function is now EXPOSED to app_main to get credentials
void tuya_get_mqtt_credentials(char* out_client_id, char* out_username, char* out_password) {
    long int t = time(NULL);
    sprintf(out_client_id, "tuyalink_%s", TUYA_DEVICE_ID);
    sprintf(out_username, "%s|signMethod=hmacSha256,timestamp=%ld,secureMode=1,accessType=1", TUYA_DEVICE_ID, t);
    String sign_content = String("deviceId=") + TUYA_DEVICE_ID + ",timestamp=" + t + ",secureMode=1,accessType=1";
    String pass_hash = hmac256(TUYA_DEVICE_SECRET, strlen(TUYA_DEVICE_SECRET), sign_content.c_str(), sign_content.length());
    strcpy(out_password, pass_hash.c_str());
}

// The new "init" function, which is non-blocking and passive.
void tuya_init(WiFiClientSecure& wifiClient, PubSubClient& mqttClient, TuyaAppCallback app_callback) {
    _mqttClient = &mqttClient;
    _app_callback = app_callback;

    // Configure the WiFiClientSecure with the necessary CA certificate
    wifiClient.setCACert(tuya_ca_cert);

    // The main app will set the server and port before connecting
}

void tuya_subscribe_topics() {
    if (_mqttClient && _mqttClient->connected()) {
        char topic_sub[128];
        sprintf(topic_sub, "tylink/%s/thing/property/set", TUYA_DEVICE_ID);
        _mqttClient->subscribe(topic_sub);
    }
}

void tuya_loop() {
    // The main app is responsible for the MQTT loop. This function does nothing.
}

bool tuya_publish_data(const String& data_json_string) {
    if (!_mqttClient || !_mqttClient->connected()) {
        return false;
    }
    char topic[128];
    sprintf(topic, "tylink/%s/thing/property/report", TUYA_DEVICE_ID);
    return _mqttClient->publish(topic, data_json_string.c_str());
}

// The internal callback that dispatches messages to the user-defined app callback
void tuya_handle_mqtt_message(char *topic, byte *payload, unsigned int length) {
    String topicStr(topic);
    // Check if this message is a Tuya message
    if (topicStr.indexOf(String("tylink/") + TUYA_DEVICE_ID) != -1) {
        String payloadStr;
        for (unsigned int i = 0; i < length; i++) {
            payloadStr += (char)payload[i];
        }
        if (_app_callback) {
            _app_callback(topicStr, payloadStr);
        }
    }
}
"""


def generate_ota_handler_header() -> str:
    return """
#ifndef OTA_HANDLER_H
#define OTA_HANDLER_H
#include <WiFi.h>
#include <PubSubClient.h>
#include <HTTPUpdate.h>

void ota_init(WiFiClient& wifiClient, PubSubClient& mqttClient, HTTPUpdate& httpUpdateClient);
void ota_loop();
const char* ota_get_device_id();

// [最终修正] 将接口升级为使用安全的String对象
void ota_handle_mqtt_message(const String& topic, const String& payload);

#endif // OTA_HANDLER_H
"""


# 文件: app/langgraph_def/graph_builder.py

def generate_ota_handler_source() -> str:
    return """
#include <WiFi.h>
#include <HTTPUpdate.h>
#include <PubSubClient.h>
#include <ArduinoJson.h>
#include "config_manager.h"
#include "ota_handler.h"
#include "mqtt_logger.h"

static WiFiClient* _wifiClient;
static PubSubClient* _mqttClient;
static HTTPUpdate* _httpUpdateClient;
static String status_topic, specific_cmd_topic, broadcast_cmd_topic;

static void perform_ota(String fileName) {
    logger.println("====== [OTA DIAGNOSIS] Step 4: perform_ota() entered. ======");
    logger.printf("Attempting OTA for file: /%s\\n", fileName.c_str());
    logger.printf("Target Server: %s:%d\\n", OTA_HTTP_SERVER, OTA_HTTP_PORT);
    if (!_wifiClient || !_httpUpdateClient) {
        logger.println("FATAL: WiFiClient or HTTPUpdateClient pointer is NULL before update!");
        return;
    }
    WiFiClient updateClient;
    logger.println("Calling httpUpdate.update()... This is a blocking call.");
    t_httpUpdate_return ret = _httpUpdateClient->update(updateClient, OTA_HTTP_SERVER, OTA_HTTP_PORT, "/" + fileName);
    logger.printf("httpUpdate.update() returned with code: %d\\n", ret);
    JsonDocument doc;
    char buffer[256];
    switch (ret) {
        case HTTP_UPDATE_FAILED:
            logger.printf("OTA RESULT: FAILED. Error (%d): %s\\n", _httpUpdateClient->getLastError(), _httpUpdateClient->getLastErrorString().c_str());
            doc["status"] = "failed";
            doc["error_code"] = _httpUpdateClient->getLastError();
            doc["error_message"] = _httpUpdateClient->getLastErrorString();
            serializeJson(doc, buffer);
            _mqttClient->publish(status_topic.c_str(), buffer);
            break;
        case HTTP_UPDATE_NO_UPDATES:
            logger.println("OTA RESULT: NO UPDATES.");
            doc["status"] = "no_update";
            serializeJson(doc, buffer);
            _mqttClient->publish(status_topic.c_str(), buffer);
            break;
        case HTTP_UPDATE_OK:
            logger.println("OTA RESULT: SUCCESS. Device will reboot.");
            doc["status"] = "success";
            serializeJson(doc, buffer);
            _mqttClient->publish(status_topic.c_str(), buffer);
            break;
    }
    logger.println("====== [OTA DIAGNOSIS] Step 5: perform_ota() finished. ======");
}

void ota_init(WiFiClient& wifiClient, PubSubClient& mqttClient, HTTPUpdate& httpUpdateClient) {
    _wifiClient = &wifiClient;
    _mqttClient = &mqttClient;
    _httpUpdateClient = &httpUpdateClient;
    status_topic = String(OTA_TOPIC_BASE) + DEVICE_ID + "/status";
    specific_cmd_topic = String(OTA_TOPIC_BASE) + DEVICE_ID + "/command";
    broadcast_cmd_topic = String(OTA_TOPIC_BASE) + "all/command";

    // [DIAGNOSTIC LOG] 确认ota_init被调用，并打印出它将要监听的主题
    logger.println("====== [OTA DIAGNOSIS] Step 0: ota_init() called. ======");
    logger.printf(" - OTA command topic set to: %s\\n", specific_cmd_topic.c_str());

    if (_mqttClient->connected()) {
        _mqttClient->subscribe(specific_cmd_topic.c_str());
        _mqttClient->subscribe(broadcast_cmd_topic.c_str());
    }
}

void ota_loop() {}
const char* ota_get_device_id() { return DEVICE_ID; }

// [最终修正] 更新函数定义，并使用安全的String对象进行操作
    void ota_handle_mqtt_message(const String& topic, const String& payload) {
        logger.println("====== [OTA DIAGNOSIS] Step 2: ota_handle_mqtt_message() entered. ======");
        
        logger.printf(" - Incoming topic: %s\\n", topic.c_str());
        logger.printf(" - Expected topic: %s\\n", specific_cmd_topic.c_str());

        if (topic != specific_cmd_topic && topic != broadcast_cmd_topic) {
            logger.println(" - Verdict: Topic MISMATCH. Exiting handler.");
            return; 
        }

        logger.println(" - Verdict: Topic MATCH. Proceeding to parse payload.");
        
        JsonDocument doc;
        DeserializationError error = deserializeJson(doc, payload);
        if (error) { 
            logger.printf("OTA JSON parsing failed: %s\\n", error.c_str()); 
            return; 
        }

        const char* action = doc["action"];
        if (action && strcmp(action, "update") == 0) {
            logger.println("====== [OTA DIAGNOSIS] Step 3: Action 'update' confirmed. ======");
            String fileName = doc["file"] | "";
            if (fileName.length() > 0) { 
                perform_ota(fileName); 
            } else {
                logger.println("ERROR: Action was 'update' but file name was missing!");
            }
        }
    }
"""

def generate_mqtt_logger_header() -> str:
    """
    V3 - 生成 MqttLogger 的头文件 (Singleton Implementation) - 已修正
    """
    return """
#ifndef MQTT_LOGGER_H
#define MQTT_LOGGER_H
#include <PubSubClient.h>
#include <Print.h>

class MqttLogger : public Print {
public:
    // 获取单例实例
    static MqttLogger& getInstance();

    // 使用MQTT客户端和设备ID初始化日志记录器
    void begin(PubSubClient& client, const char* device_id);

    // 在主循环中调用此函数以处理缓冲区刷新
    void loop();

    // Print 接口方法
    virtual size_t write(uint8_t);
    virtual size_t write(const uint8_t *buffer, size_t size);

private:
    // 私有构造函数以强制执行单例模式
    MqttLogger();

    // 私有析构函数
    ~MqttLogger();

    // 删除拷贝构造函数和赋值运算符
    MqttLogger(const MqttLogger&) = delete;
    MqttLogger& operator=(const MqttLogger&) = delete;

    PubSubClient* _client;
    String _topic;
    char _buffer[256];
    size_t _buffer_pos;
    unsigned long _last_flush;
    bool _initialized;

    void flush();
};

// V3.2 修正: 添加 extern 声明，让所有包含此头文件的文件都知道 logger 的存在。
// 它的实体将在 app_main.ino 中被定义。
extern Print& logger;

#endif // MQTT_LOGGER_H
"""

def generate_mqtt_logger_source() -> str:
    """
    V3 - 生成 MqttLogger 的源文件 (Singleton Implementation)
    """
    return """
#include "mqtt_logger.h"
#include "config_manager.h" // For DEBUG_TOPIC_BASE

// 唯一的静态实例
MqttLogger& MqttLogger::getInstance() {
    static MqttLogger instance;
    return instance;
}

// 私有构造函数
MqttLogger::MqttLogger()
    : _client(nullptr), _buffer_pos(0), _last_flush(0), _initialized(false) {}

// 私有析构函数
MqttLogger::~MqttLogger() {}

// 初始化方法
void MqttLogger::begin(PubSubClient& client, const char* device_id) {
    _client = &client;
    _topic = String(DEBUG_TOPIC_BASE) + device_id + "/log";
    _initialized = true;
}

void MqttLogger::loop() {
    if (_initialized && _buffer_pos > 0 && (millis() - _last_flush > 1000)) {
        flush();
    }
}

size_t MqttLogger::write(uint8_t c) {
    if (!_initialized) return 0;
    if (_buffer_pos >= sizeof(_buffer) - 1) {
        flush();
    }
    _buffer[_buffer_pos++] = c;
    if (c == '\\n') { // 在换行时刷新以获得更好的响应性
        flush();
    }
    return 1;
}

size_t MqttLogger::write(const uint8_t *buffer, size_t size) {
    if (!_initialized) return 0;
    for (size_t i = 0; i < size; i++) {
        write(buffer[i]);
    }
    return size;
}

void MqttLogger::flush() {
    if (_initialized && _buffer_pos > 0 && _client && _client->connected()) {
        _buffer[_buffer_pos] = '\\0';
        _client->publish(_topic.c_str(), _buffer);
        _buffer_pos = 0;
    }
    _last_flush = millis();
}
"""

# =================================================================================
# 4. Agent Node Definitions (无变动)
# =================================================================================

def module_architect_node(state: AgentState) -> Dict:
    device_task = state['current_device_task']
    device_id = device_task['internal_device_id']
    print(f"--- L2: MODULE ARCHITECT: Designing firmware for '{device_id}' ---")

    peripherals_info_parts = []
    for p in device_task.get('peripherals', []):
        model_str = f" (Model: {p['model']})" if p.get('model') and p['model'].lower() not in ['generic', ''] else ""
        peripherals_info_parts.append(f"- {p['name']}{model_str}")
    peripherals_info = "\\n".join(peripherals_info_parts)

    prompt = textwrap.dedent(f"""
    <Prompt>
        <Role>You are an expert embedded firmware architect.</Role>
        <Goal>For the given device, design a modular firmware architecture that includes all necessary functionalities.</Goal>
        <Context>
            <Device>{device_id} ({device_task['board']})</Device>
            <DeviceRole>{device_task['description']}</DeviceRole>
            <Peripherals>
            {peripherals_info}
            </Peripherals>
        </Context>
        <Instructions>
        1.  **Driver Modules**: For each physical peripheral, define a 'driver' module. The `task_id` should be based on the peripheral's Model or Name, ending with `_driver`.
        2.  **Application Module**: Define one single 'application' module named `app_main` that uses all other modules.
        3.  **Core Services (Mandatory)**: You MUST ALWAYS include these three core service modules: `config_manager`, `ota_handler`, and `mqtt_logger`. They are essential for any network-connected device.
        4.  **Cloud Service (Conditional)**: If the device's description or peripherals explicitly mention "Tuya", you MUST ADDITIONALLY include the `tuya_handler` module.
        5.  **Dependencies**: The `app_main` module MUST list all other generated modules as its dependencies.
        6.  **Output Format**: Your final output MUST be a single, valid JSON object containing one key: "modules".
        7.  **No Peripherals Rule**: If a device has no physical peripherals listed in its `<Peripherals>` context, you MUST NOT generate any hardware-specific driver modules for it (e.g., no `light_sensor_driver`). The only modules should be the mandatory Core Services, an optional Cloud Service, and the Application module.
        </instructions>
        <ExampleOutput>
        ```json
        {{
            "modules": [
                {{ "task_id": "config_manager", "task_type": "driver", "peripheral": "Core", "description": "Manages all network and device configurations.", "dependencies": [] }},
                {{ "task_id": "ota_handler", "task_type": "driver", "peripheral": "Core", "description": "Handles over-the-air firmware updates.", "dependencies": [] }},
                {{ "task_id": "mqtt_logger", "task_type": "driver", "peripheral": "Core", "description": "Handles remote logging over MQTT for debugging and verification.", "dependencies": [] }},
                {{ "task_id": "tuya_handler", "task_type": "driver", "peripheral": "Core", "description": "Handles connection and data exchange with the Tuya Cloud.", "dependencies": [] }},
                {{ "task_id": "bh1750_driver", "task_type": "driver", "peripheral": "BH1750", "description": "A driver for the BH1750 light sensor.", "dependencies": [] }},
                {{ "task_id": "app_main", "task_type": "application", "description": "The main application logic.", "dependencies": ["config_manager", "ota_handler", "mqtt_logger", "tuya_handler", "bh1750_driver"] }}
            ]
        }}
        ```
        </ExampleOutput>
    </Prompt>
    """)
    response = module_architect_model.invoke([HumanMessage(content=prompt)])
    try:
        plan = json.loads(extract_code(response.content, "json"))
        return {"module_tasks": plan['modules'], "original_module_plan": plan['modules']}
    except (json.JSONDecodeError, KeyError, TypeError) as e:
        print(f"MODULE ARCHITECT PARSING ERROR: {e}")
        return {"feedback": f"FAIL: Module Architect failed. Error: {e}", "module_tasks": []}


def plan_enrichment_node(state: AgentState) -> Dict:
    """
    [V3 - Refined]: An AI node that refines user-friendly requirements into machine-friendly instructions in English.
    It now generates specific, structured MQTT topics.
    """
    print("--- [PLAN ENRICHMENT V3]: Refining device task descriptions into English and generating MQTT topics... ---")

    project_name = state.get('project_name', 'default_project')
    device_tasks = state.get('device_tasks_queue', [])
    communication_plan = state.get('system_plan', {}).get('communication', [])

    if not device_tasks:
        return {}

    # 1. Create communication map with structured topics
    topic_map = {}
    # Sanitize project name for use in topic
    safe_project_name = re.sub(r'[^a-zA-Z0-9_-]', '', project_name.lower().replace(' ', '_'))

    for comm in communication_plan:
        source_role = comm.get('source_device_role')
        target_role = comm.get('target_device_role')

        safe_source_role = re.sub(r'[^a-zA-Z0-9_-]', '', source_role.lower().replace(' ', '_'))
        topic = f"/{safe_project_name}/{safe_source_role}/data"

        if source_role not in topic_map: topic_map[source_role] = {"pub": [], "sub": []}
        if target_role not in topic_map: topic_map[target_role] = {"pub": [], "sub": []}

        if topic not in topic_map[source_role]["pub"]:
            topic_map[source_role]["pub"].append(topic)
        if topic not in topic_map[target_role]["sub"]:
            topic_map[target_role]["sub"].append(topic)

    # 2. Iterate through each device task and call LLM to rewrite
    for task in device_tasks:
        role = task.get('device_role')
        original_description = task.get('description', '')

        comm_context = "This device has no assigned inter-device communication tasks."
        if role in topic_map:
            pub_info = f"It MUST publish data to the following MQTT topics: {', '.join(topic_map[role]['pub'])}." if \
            topic_map[role]["pub"] else ""
            sub_info = f"It MUST subscribe to the following MQTT topics to receive data: {', '.join(topic_map[role]['sub'])}." if \
            topic_map[role]["sub"] else ""
            comm_context = " ".join(filter(None, [pub_info, sub_info]))

        rewrite_prompt = textwrap.dedent(f"""
        <Prompt>
            <Role>You are a technical writer specializing in embedded systems. Your task is to rewrite a high-level functional description into a precise, unambiguous technical specification in English for a developer.</Role>
            <Goal>Combine the device's core function with its specific communication tasks into a single, clear paragraph in English.</Goal>
            <Context>
                <DeviceRole>{role}</DeviceRole>
                <OriginalDescription>{original_description}</OriginalDescription>
                <TechnicalCommunicationPlan>{comm_context}</TechnicalCommunicationPlan>
            </Context>
            <Instructions>
                1.  Translate the core function from the `<OriginalDescription>` into clear, concise English.
                2.  Integrate the specific actions from the `<TechnicalCommunicationPlan>`. Replace vague phrases like "send data to another device" with explicit technical actions like "publish data to the MQTT topic ...".
                3.  The final output MUST be a single, concise paragraph in English.
                4.  Do not add any new functionality. Your job is to translate and specify, not invent.
            </Instructions>
            <Example>
                <Input>
                    <OriginalDescription>使用BH1750传感器读取光照强度，并将数据发送给小黑板和涂鸦云。</OriginalDescription>
                    <TechnicalCommunicationPlan>It MUST publish data to the following MQTT topics: /smart_light_system/light_sensor/data.</TechnicalCommunicationPlan>
                </Input>
                <Output>
                Periodically read the light intensity from the BH1750 sensor. Publish the illumination data to the MQTT topic /smart_light_system/light_sensor/data for inter-device communication. Simultaneously, report the same data to the Tuya Cloud.
                </Output>
            </Example>
            <FinalOutput>
            {{Your rewritten English description here as a single string.}}
            </FinalOutput>
        </Prompt>
        """)

        response = system_architect_model.invoke([HumanMessage(content=rewrite_prompt)])
        new_description = response.content.strip()

        task['description'] = new_description
        print(f"  -> Rewrote description for '{role}':")
        print(f"     [Before]: {original_description}")
        print(f"     [After]:  {task['description']}")

    return {"device_tasks_queue": device_tasks}

def device_dispatcher_node(state: AgentState) -> Dict:
    queue = state.get('device_tasks_queue', [])
    if not queue:
        # 所有设备处理完毕，返回一个特殊的None值，用于路由到最终的文件生成节点
        return {"current_device_task": None}

    # 准备处理下一个设备
    next_device_task = queue[0]
    remaining_queue = queue[1:]

    # --- [最终修正] 为下一个设备循环创建一个全新的、干净的返回字典 ---

    # 1. 保留需要在设备间传递的持久化状态
    persistent_state = {
        "user_input": state.get('user_input'),
        "system_plan": state.get('system_plan'),
        "user_id": state.get('user_id'),
        "workflow_id": state.get('workflow_id'),
        "workspace_path": state.get('workspace_path'),
        "project_name": state.get('project_name'),
        "dp_info_list": state.get('dp_info_list', []),  # 累积DP列表
        "project_files": state.get('project_files', {})  # 累积项目文件
    }

    # 2. 重置UI步骤状态
    current_steps = state.get('workflow_steps', [])
    dev_loop_steps = [
        "module_architect", "module_dispatcher", "api_designer", "developer",
        "integrator", "test_plan_designer", "deployment_and_verification",
        "compile_node", "pre_deployment_pause", "usb_upload_node",
        "ota_deployment_node", "deploy_and_verify_node", "dp_extractor"
    ]
    for step in current_steps:
        if step['id'] in dev_loop_steps:
            step['status'] = 'pending'
            # 保留历史日志，不清空 step['log']
            step['start_time'] = 0.0
            step['end_time'] = 0.0
            step['output'] = None

    # 3. 从数据库加载新设备所需的信息
    user = User.query.get(state['user_id'])
    device = Device.query.filter_by(internal_device_id=next_device_task['internal_device_id']).first()

    # 4. 构建最终的返回字典
    update_dict = {
        **persistent_state,
        "workflow_steps": current_steps,  # 使用我们刚刚重置过的列表
        "device_tasks_queue": remaining_queue,
        "current_device_task": next_device_task,
        # 显式重置所有临时开发状态
        "module_tasks": [],
        "current_module_task": None,
        "completed_modules": {},
        "feedback": "",
        "original_module_plan": None,
        "current_api_spec": None,
        "test_plan": None,
        "build_dir": "",
        "firmware_path": None,
    }

    # 填充数据库信息到新状态中
    if user:
        print(f"  -> Found User '{user.username}'. Loading WiFi credentials.")
        update_dict['wifi_ssid'] = user.wifi_ssid
        update_dict['wifi_password'] = user.wifi_password

    if device:
        print(f"  -> Found Device '{device.nickname}'. Loading Tuya credentials.")
        update_dict['cloud_product_id'] = device.cloud_product_id
        update_dict['cloud_device_id'] = device.cloud_device_id
        update_dict['cloud_device_secret'] = device.cloud_device_secret

    return update_dict


def module_dispatcher_node(state: AgentState) -> Dict:
    if state.get('module_tasks'):
        next_task = state['module_tasks'][0]
        print(f"--- L3 DISPATCHER: Selecting module task -> '{next_task['task_id']}' ---")
        return {"current_module_task": next_task, "module_tasks": state['module_tasks'][1:], "feedback": ""}
    return {"current_module_task": None}


def api_designer_node(state: AgentState) -> Dict:
    """
    通过“本地缓存 -> 本地模糊搜索 -> AI生成”三段式流程获取API。
    """
    task = state['current_module_task']
    if not task or task['task_type'] != 'driver' or task['task_id'] in ['config_manager', 'ota_handler', 'mqtt_logger',
                                                                        'tuya_handler']:
        return {"current_api_spec": None}

    peripheral = task['peripheral']
    print(f"--- L3: API DESIGNER: Searching for API for '{peripheral}' ---")

    # 阶段 1 & 2: 本地精确/模糊搜索
    cached_spec_str = find_api_spec(peripheral)
    if cached_spec_str:
        try:
            spec_json = json.loads(cached_spec_str)
            interface = None
            if isinstance(spec_json, dict):
                for key, value in spec_json.items():
                    if isinstance(value, dict) and 'functions' in value:
                        interface = value
                        break

            if interface:
                functions = interface.get("functions", [])
                formatted_spec = json.dumps(functions, indent=2, ensure_ascii=False)
                print(f"--- API DESIGNER: Successfully loaded API for '{peripheral}' from local cache. ---")
                return {"current_api_spec": formatted_spec}
            else:
                print(
                    f"--- API DESIGNER: Found local file for '{peripheral}', but content format is unexpected. Proceeding to AI generation. ---")
        except Exception as e:
            print(
                f"--- API DESIGNER: Error parsing local API file for '{peripheral}': {e}. Proceeding to AI generation. ---")

    # 阶段 3: AI 生成 (作为备选)
    print(f"--- L3: API DESIGNER: No local API found. Generating with AI for '{peripheral}' ---")
    prompt = textwrap.dedent(f"""
    <Prompt>
        <Role>You are an expert API designer for embedded C/C++ drivers.</Role>
        <Goal>Generate a high-quality, detailed API specification in JSON format for the given peripheral or logical module.</Goal>
        <Context>
            <PeripheralOrModule>{peripheral}</PeripheralOrModule>
            <Task>{task['description']}</Task>
        </Context>
        <Instructions>
            1. Design a set of C-style functions.
            2. For communication modules like MQTT, design high-level functions like `connect`, `publish`, `subscribe`.
            3. The output must be a single, valid JSON object containing a root key `"{peripheral.upper().replace(' ', '_')}_Interface"` which contains a list of functions.
        </Instructions>
        <Example>
        For a 'DHT11', a good output would be:
        ```json
        {{
            "DHT11_Interface": {{
                "functions": [
                    {{ "name": "dht11_setup", "description": "Initializes the DHT11 sensor on a specific pin.", "return_type": "void", "parameters": [{{"name": "pin", "type": "int"}}] }},
                    {{ "name": "dht11_read_temperature", "description": "Reads the temperature in Celsius.", "return_type": "float", "parameters": [] }},
                    {{ "name": "dht11_read_humidity", "description": "Reads the humidity in percent.", "return_type": "float", "parameters": [] }}
                ]
            }}
        }}
        ```
        </Example>
        <OutputFormat>```json
// Your generated JSON code here
```</OutputFormat>
    </Prompt>
    """)
    response = api_designer_model.invoke([HumanMessage(content=prompt)])
    generated_spec_str = extract_code(response.content, lang="json")
    try:
        spec_json = json.loads(generated_spec_str)
        interface_key_from_llm = f"{peripheral.upper().replace(' ', '_')}_Interface"
        interface = spec_json.get(interface_key_from_llm, {})
        functions = interface.get("functions", [])
        formatted_spec = json.dumps(functions, indent=2, ensure_ascii=False)
        print(f"--- API DESIGNER: Successfully generated API for '{peripheral}' using AI. ---")

        try:
            save_content = {interface_key_from_llm: {"functions": functions}}
            save_filename = f"{peripheral.upper().replace(' ', '_')}_API_Package.json"
            save_filepath = API_PACKAGE_DIR / save_filename
            if not API_PACKAGE_DIR.exists():
                API_PACKAGE_DIR.mkdir(parents=True, exist_ok=True)
            save_filepath.write_text(json.dumps(save_content, indent=4, ensure_ascii=False), encoding='utf-8')
            print(f"--- API DESIGNER: Saved newly generated API to local cache: '{save_filename}' ---")
        except Exception as e:
            print(f"--- API DESIGNER: WARNING - Failed to save newly generated API to cache: {e} ---")

        return {"current_api_spec": formatted_spec}
    except Exception as e:
        print(f"API DESIGNER GENERATION ERROR: {e}")
        return {"current_api_spec": f"// Failed to generate API for {peripheral}"}


def developer_node(state: AgentState) -> Dict:
    feedback = state.get('feedback', '')
    task = state.get('current_module_task')

    if not task and "FAIL" in feedback:
        print("--- [DEVELOPER] Received compilation failure, attempting to identify faulty module for repair. ---")
        faulty_module_id = None
        primary_pattern = re.compile(r'^(.*?[/\\](src|lib)[/\\][^/\\:]+\.cpp):\d+:\d+:\s+error:', re.MULTILINE)
        match = primary_pattern.search(feedback)

        if match:
            filepath = Path(match.group(1))
            faulty_module_id = filepath.stem
            print(f"--- [DEVELOPER] Inferred faulty module from error line: '{faulty_module_id}' ---")
        else:
            fallback_pattern = re.compile(r"\[\.pio[/\\]build[/\\].*?[/\\](src|lib)[/\\].*?([^\s/\\]+)\.cpp\.o\] Error",
                                          re.MULTILINE)
            match = fallback_pattern.search(feedback)
            if match:
                faulty_module_id_with_ext = match.group(2)
                if '.ino.cpp' in faulty_module_id_with_ext:
                    faulty_module_id = faulty_module_id_with_ext.replace('.ino.cpp', '')
                elif '.cpp' in faulty_module_id_with_ext:
                    faulty_module_id = faulty_module_id_with_ext.replace('.cpp', '')
                elif '.ino' in faulty_module_id_with_ext:
                    faulty_module_id = faulty_module_id_with_ext.replace('.ino', '')
                else:
                    faulty_module_id = faulty_module_id_with_ext
                print(
                    f"--- [DEVELOPER] Inferred faulty module using fallback summary pattern: '{faulty_module_id_with_ext}' -> Cleaned to '{faulty_module_id}' ---")

        if not faulty_module_id:
            faulty_module_id = "app_main"
            print(f"--- [DEVELOPER] Could not infer module from logs. Defaulting to '{faulty_module_id}'. ---")

        original_plan = state.get('original_module_plan', [])
        faulty_task_details = next((m for m in original_plan if m['task_id'] == faulty_module_id), None)

        if faulty_task_details:
            print(f"--- [DEVELOPER] Identified '{faulty_module_id}' as the module to repair. Retrying development. ---")
            task = faulty_task_details
        else:
            print(
                f"--- [DEVELOPER] Could not find details for module '{faulty_module_id}' in original plan. Aborting repair. ---")
            return {"feedback": f"FAIL: Could not identify faulty module '{faulty_module_id}' for repair."}

    if not task:
        return {}

    device_id = state['current_device_task']['internal_device_id']
    task_id = task['task_id']
    print(f"--- L3: DEVELOPER: Coding module '{task_id}' for device '{device_id}' ---")

    feedback_context = ""
    if "FAIL" in feedback:
        # 核心修正：截断过长的错误日志，只保留最后8000个字符
        MAX_FEEDBACK_LENGTH = 8000
        trimmed_feedback = feedback
        if len(trimmed_feedback) > MAX_FEEDBACK_LENGTH:
            print(
                f"  [INFO] Feedback log is too long ({len(trimmed_feedback)} chars). Truncating to last {MAX_FEEDBACK_LENGTH} chars.")
            trimmed_feedback = feedback[-MAX_FEEDBACK_LENGTH:]

        feedback_context = textwrap.dedent(f"""
        <Feedback_From_Previous_Attempt>
        IMPORTANT: Your previous attempt to generate code for this module resulted in a failure. You MUST analyze the following error message and fix the code accordingly.
        Error Details: {trimmed_feedback}
        Common root causes for 'conflicting declaration' errors are due to incorrect singleton implementation or multiple definitions of a global variable. Ensure the logger is defined once as a static instance and accessed via a public static method.
        Another common error is 'was not declared in this scope', which means you forgot to `#include` the header file for a function you are calling.
        You MUST ensure your generated code is complete and correct.
        </Feedback_From_Previous_Attempt>
        """)

    completed_modules = state.get('completed_modules', {})
    version = completed_modules.get(task_id, {}).get('version', 0) + 1

    if task_id == 'config_manager':
        header_code = generate_config_manager_header(
            device_id=state['current_device_task']['internal_device_id'],
            wifi_ssid=state.get('wifi_ssid'),
            wifi_password=state.get('wifi_password'),
            product_id=state.get('cloud_product_id'),
            cloud_device_id=state.get('cloud_device_id'),
            device_secret=state.get('cloud_device_secret')
        )
        completed_modules[task_id] = {"task_id": task_id, "header_code": header_code, "source_code": None,
                                      "main_code": None, "version": version}
        return {"completed_modules": completed_modules, "feedback": "", "current_module_task": task}

    if task_id == 'ota_handler':
        header_code = generate_ota_handler_header()
        source_code = generate_ota_handler_source()
        completed_modules[task_id] = {"task_id": task_id, "header_code": header_code, "source_code": source_code,
                                      "main_code": None, "version": version}
        return {"completed_modules": completed_modules, "feedback": "", "current_module_task": task}

    if task_id == 'mqtt_logger':
        header_code = generate_mqtt_logger_header()
        source_code = generate_mqtt_logger_source()
        completed_modules[task_id] = {"task_id": task_id, "header_code": header_code, "source_code": source_code,
                                      "main_code": None, "version": version}
        return {"completed_modules": completed_modules, "feedback": "", "current_module_task": task}

    if task_id == 'tuya_handler':
        header_code = generate_tuya_handler_header()
        source_code = generate_tuya_handler_source()
        sha256_h_content = generate_mbedtls_sha256_header()
        sha256_cpp_content = generate_mbedtls_sha256_source()
        source_files_dict = {"tuya_handler.cpp": source_code, "SHA256.h": sha256_h_content,
                             "SHA256.cpp": sha256_cpp_content}
        completed_modules[task_id] = {"task_id": task_id, "header_code": header_code,
                                      "source_code": json.dumps(source_files_dict), "main_code": None,
                                      "version": version}
        return {"completed_modules": completed_modules, "feedback": "", "current_module_task": task}

    api_spec = state.get('current_api_spec', 'No API specification provided.')
    context, instructions = "", ""
    if task['task_type'] == 'driver':
        context = f"<APISpecification>\\n{api_spec}\\n</APISpecification>"
        instructions = f"""
                            <Instructions>
                            1.  **Goal**: Your task is to implement the C++ header (.h) and source (.cpp) files for the driver based *only* on the provided `<APISpecification>`.
                            2.  **Strict Adherence**: You MUST implement every function exactly as defined in the spec. Do not add, remove, or modify any functions.
                            3.  **Completeness**: Provide the complete code for both the header and the source file. Do not omit any part.
                            4.  **Header File Guard**: The header file MUST include standard header guards (`#ifndef`, `#define`, `#endif`).
                            5.  **ESP32 I2C Driver**: If implementing an I2C peripheral for ESP32, you MUST use the ESP-IDF driver functions from `<driver/i2c.h>`.
                            6.  For `i2c_master_read`, the final read operation requires `I2C_MASTER_NACK` as the last argument to properly terminate the communication.
                            7.  Do NOT use `0` or `I2C_MASTER_ACK` for the final read.
                            8.  **No `main()`**: Do not include a `main()` function or `setup()`/`loop()` unless the API spec explicitly requires it. These are library files.
                            9.  **No Placeholders**: Your code must be fully implemented and functional. Do not leave placeholder comments like `// Your implementation here`.
                            </Instructions>
                            <CorrectExample for ESP32 I2C Driver>
                           [HEADER]
                           ```cpp
                           #ifndef SIMPLE_I2C_DRIVER_H
                           #define SIMPLE_I2C_DRIVER_H
                           #include "driver/i2c.h"

                           void simple_i2c_init(i2c_port_t port, int sda_pin, int scl_pin);
                           esp_err_t simple_i2c_read_byte(i2c_port_t port, uint8_t device_addr, uint8_t reg_addr, uint8_t *data);
                           #endif
                           ```

                           [SOURCE]
                           ```cpp
                           #include "simple_i2c_driver.h"

                           void simple_i2c_init(i2c_port_t port, int sda_pin, int scl_pin) {{{{
                               i2c_config_t conf;
                               conf.mode = I2C_MODE_MASTER;
                               conf.sda_io_num = sda_pin;
                               conf.scl_io_num = scl_pin;
                               conf.sda_pullup_en = GPIO_PULLUP_ENABLE;
                               conf.scl_pullup_en = GPIO_PULLUP_ENABLE;
                               conf.master.clk_speed = 100000;
                               i2c_param_config(port, &conf);
                               i2c_driver_install(port, conf.mode, 0, 0, 0);
                           }}}}

                           esp_err_t simple_i2c_read_byte(i2c_port_t port, uint8_t device_addr, uint8_t reg_addr, uint8_t *data) {{{{
                               return ESP_OK;
                           }}}}
                           ```
                           </CorrectExample>
                            <OutputFormat>
                           You MUST provide two distinct code blocks, one for the header and one for the source file. Use the specified markdown format.
                            [HEADER]
                            ```cpp
                            // Header file content for {{{{task['task_id']}}}}.h
                            ```
                           [SOURCE]
                            ```cpp
                            // Source file content for {{{{task['task_id']}}}}.cpp
                            ```
                            </OutputFormat>
                            """
        prompt = textwrap.dedent(f"""
        <Prompt>
            <Role>You are an expert embedded systems developer following a strict modular architecture.</Role>
            <Context>
                <TaskDescription>{task['description']}</TaskDescription>
                {feedback_context}
                {context}
            </Context>
            {instructions}
        </Prompt>
        """)
        response = developer_model.invoke([HumanMessage(content=prompt)])
        content = response.content
        header_code = extract_code(content, lang="cpp", block_name="HEADER")
        source_code = extract_code(content, lang="cpp", block_name="SOURCE")
        completed_modules[task_id] = {"task_id": task_id, "header_code": header_code, "source_code": source_code,
                                      "main_code": None, "version": version}
    else:  # application
        print("--- L3: DEVELOPER: Coding main application 'app_main.ino' ---")
        driver_headers = ""
        completed = state.get('completed_modules', {})
        dependencies = task.get('dependencies', [])

        for dep_id in dependencies:
            if dep_id in completed and completed[dep_id].get('header_code'):
                driver_headers += f"--- Interface for {dep_id} from '{dep_id}.h' ---\\n```cpp\\n{completed[dep_id]['header_code']}\\n```\\n\\n"

        is_tuya_device = "tuya_handler" in dependencies

        if is_tuya_device:
            print("  -> Generating DUAL-CLIENT (Tuya + Local) architecture for app_main.ino")
            prompt = textwrap.dedent(f"""
                <Prompt>
                    <Role>You are an expert embedded firmware developer for the ESP32.</Role>
                    <Goal>Generate the complete `app_main.ino` file for a dual-client MQTT application.</Goal>
                    <Architectural_Mandates>
                        1.  **Strict Task Focus**: Your ONLY task is to implement the logic described in `<TaskDescription>`.
                        2.  **Implement ALL Communications**: The `<TaskDescription>` may contain instructions for both public cloud communication and local inter-device communication. Your code in `loop()` MUST implement **ALL** specified communication paths.
                        3.  **No Extra Logic**: Implement the logic exactly as described. Do NOT add any extra conditions or thresholds that were not specified.
                        4.  **Error Handling**: Before publishing sensor data, you MUST check if the reading is valid. For many sensors, a return value of -1 or less indicates an error. If an error is detected, you should log the error (e.g., `logger.println("Failed to read sensor.");`) and skip publishing for that cycle.
                        5.  **Correct Tuya Payload**: When using `tuya_publish_data`, the payload string MUST be a valid JSON formatted as `"{{\\"properties\\":{{\\"your_dp_code\\":...}}}}"`.
                        6.  **Include Headers**: You MUST include the necessary header files for all driver modules used, as detailed in `<DriverInterfaces>`.
                        7.  **Confirmation Logging**: Immediately after any successful `localMqttClient.publish()` or `tuya_publish_data()` call, you MUST add a log line confirming the action, for example: `logger.println("Published data to Tuya cloud.");`. This is mandatory for verification.
                        8.  **Polling Interval**: The main `loop()` function MUST include a non-blocking delay mechanism to ensure it runs at a reasonable interval (e.g., every 5-10 seconds). Use a `static unsigned long ...` and `if (millis() - ... > ...)` pattern.
                        9.  **Dual-Client Architecture & RTOS**: You MUST use two separate `PubSubClient` instances and a dedicated FreeRTOS task (`tuyaConnectionTask`) for the Tuya connection, as shown in the example.
                    </Architectural_Mandates>
                    <Context>
                        <TaskDescription>{task['description']}</TaskDescription>
                        <DriverInterfaces>{driver_headers}</DriverInterfaces>
                        <IsTuyaDevice>{is_tuya_device}</IsTuyaDevice>
                        {feedback_context}
                    </Context>
                    <CorrectExample_for_Tuya_Dual_Client_Architecture>
                    ```cpp
                            // --- 核心库 ---
                            #include <WiFi.h>
                            #include <PubSubClient.h>
                            #include <ArduinoJson.h>
                            #include <WiFiClientSecure.h>
                            #include "time.h"
        
                            // --- 项目中的功能模块库 ---
                            #include "config_manager.h"
                            #include "tuya_handler.h"
                            #include "ota_handler.h"
                            #include "mqtt_logger.h"
                            #include <HTTPUpdate.h>
                            // #include "bh1750_driver.h" // EXAMPLE DRIVER - You must include the actual drivers you use.
        
                            // =======================================================================
                            // 1. 双客户端定义
                            // =======================================================================
                            WiFiClientSecure tuyaWifiClient;
                            PubSubClient tuyaMqttClient(tuyaWifiClient);
                            WiFiClient localWifiClient;
                            HTTPUpdate myHttpUpdate;
                            PubSubClient localMqttClient(localWifiClient);
                            Print& logger = MqttLogger::getInstance();
        
                            // =======================================================================
                            // 2. 回调函数定义
                            // =======================================================================
                            void handle_tuya_app_commands(String &topic, String &payload) {{
                                logger.println("Received application command from Tuya Cloud via handler.");
                                // Add logic here to handle commands like `enable_report`
                            }}
        
                            void tuya_mqtt_callback(char* topic, byte* payload, unsigned int length) {{
                                tuya_handle_mqtt_message(topic, payload, length);
                            }}
        
                            void local_mqtt_callback(char* topic, byte* payload, unsigned int length) {{
                                String topicStr = String(topic);
                                String payloadStr;
                                payloadStr.reserve(length);
                                for (unsigned int i = 0; i < length; i++) {{ payloadStr += (char)payload[i]; }}
                                ota_handle_mqtt_message(topicStr, payloadStr);
                            }}
        
                            // =======================================================================
                            // 3. 连接函数
                            // =======================================================================
                            void connectToTuya() {{
                                logger.println("Attempting to connect to TUYA MQTT Broker...");
                                char clientId[128], username[256], password[128];
                                tuya_get_mqtt_credentials(clientId, username, password);
                                if (tuyaMqttClient.connect(clientId, username, password)) {{
                                    logger.println("SUCCESS: Connected to Tuya MQTT Broker.");
                                    tuya_subscribe_topics();
                                }} else {{
                                    logger.printf("FAILED, Tuya client state=%d.\\n", tuyaMqttClient.state());
                                }}
                            }}
        
                            void tuyaConnectionTask(void *pvParameters) {{
                                logger.println("Tuya Connection Task started.");
                                for (;;) {{
                                    if (!tuyaMqttClient.connected()) {{
                                        connectToTuya();
                                        if (!tuyaMqttClient.connected()) {{
                                            vTaskDelay(5000 / portTICK_PERIOD_MS);
                                        }}
                                    }}
                                    tuyaMqttClient.loop();
                                    vTaskDelay(20 / portTICK_PERIOD_MS);
                                }}
                            }}
        
                            void connectToLocalMqtt() {{
                                while (!localMqttClient.connected()) {{
                                    logger.println("Attempting to connect to LOCAL MQTT Broker...");
                                    if (localMqttClient.connect(DEVICE_ID)) {{
                                        logger.println("SUCCESS: Connected to Local MQTT Broker.");
                                        ota_init(localWifiClient, localMqttClient, myHttpUpdate);
                                    }} else {{
                                        logger.printf("FAILED, rc=%d. Retrying in 5 seconds\\n", localMqttClient.state());
                                        delay(5000);
                                    }}
                                }}
                            }}
        
                            // =======================================================================
                            // 4. 主程序: Setup & Loop
                            // =======================================================================
                            void setup() {{
                                Serial.begin(115200);
                                delay(10);
                                Serial.println("\\n--- Dual-Client Firmware with Dedicated Task ---");
                                WiFi.begin(WIFI_SSID, WIFI_PASSWORD);
                                while (WiFi.status() != WL_CONNECTED) {{ delay(500); Serial.print("."); }}
                                Serial.println("\\nWiFi connected.");
                                configTime(0, 0, "pool.ntp.org", "time.nist.gov");
                                time_t now = time(NULL);
                                while (now < 8 * 3600 * 2) {{ delay(500); now = time(NULL); }}
                                MqttLogger::getInstance().begin(localMqttClient, DEVICE_ID);
                                logger.println("--- System Initializing ---");
        
                                // Initialize handlers
                                tuya_init(tuyaWifiClient, tuyaMqttClient, handle_tuya_app_commands);
        
                                // Configure MQTT clients
                                tuyaMqttClient.setServer("m1.tuyacn.com", 8883);
                                tuyaMqttClient.setCallback(tuya_mqtt_callback);
                                localMqttClient.setServer(MQTT_BROKER, MQTT_PORT);
                                localMqttClient.setCallback(local_mqtt_callback);
        
                                // Connect to local broker and initialize OTA
                                connectToLocalMqtt();
        
                                // Create Tuya connection task
                                xTaskCreate(tuyaConnectionTask, "TuyaTask", 10240, NULL, 1, NULL);
        
                                logger.println("--- Setup Complete, Main Loop is starting ---");
                            }}
                            void loop() {{
                                if (!localMqttClient.connected()) {{ connectToLocalMqtt(); }}
                                localMqttClient.loop();
                                MqttLogger::getInstance().loop();
                                // Add application logic here, e.g., reading sensors and publishing data.
                            }}
                            ```
                            </CorrectExample_for_Tuya_Dual_Client_Architecture>
                            <OutputFormat>You MUST provide a single, complete code block for the main `.ino` file. Your code must strictly follow all rules and the style of the provided example.</OutputFormat>
                        </Prompt>
                    """)
        else:
            print("  -> Generating SINGLE-CLIENT (Local only) architecture for app_main.ino")
            prompt = textwrap.dedent(f"""
                            <Prompt>
                                <Role>You are an expert embedded firmware developer creating a robust IoT application for the ESP32.</Role>
                                <Goal>Your primary task is to generate the *entire* `app_main.ino` file. The firmware only needs to connect to a **local MQTT broker** for debugging and OTA updates.</Goal>
                                <Architectural_Mandates>
                                    1.  **Strict Task Focus**: Your ONLY task is to implement the logic described in `<TaskDescription>`.
                                    2.  **No Extra Logic**: Implement the logic exactly as described. Do NOT add any extra conditions (like `if` statements) or thresholds that were not specified.
                                    3.  **Include Headers**: You MUST include the necessary header files for all driver modules used, as detailed in `<DriverInterfaces>`.
                                    4.  **Confirmation Logging**: Immediately after any successful `localMqttClient.publish()` call, you MUST add a log line confirming the action, for example: `logger.println("Published data to local MQTT.");`. This is mandatory for verification.
                                    5.  **Polling Interval**: The main `loop()` function MUST include a non-blocking delay mechanism to ensure it runs at a reasonable interval (e.g., every 5-10 seconds). Use a `static unsigned long lastActionTime = 0;` and `if (millis() - lastActionTime > 5000) {{ ... }}` pattern.
                                    6.  **Single-Client Architecture**: You MUST define only ONE `PubSubClient` instance: `localMqttClient` (using `WiFiClient`).
                                </Architectural_Mandates>
                                <Context>
                                    <TaskDescription>{task['description']}</TaskDescription>
                                    <DriverInterfaces>{driver_headers}</DriverInterfaces>
                                    <IsTuyaDevice>{is_tuya_device}</IsTuyaDevice>
                                    {feedback_context}
                                </Context>
                                <CorrectExample_for_Single_Client_Architecture>
                                ```cpp
                                // --- Core Libraries ---
                                #include <WiFi.h>
                                #include <PubSubClient.h>
                                #include <ArduinoJson.h>
                                #include <HTTPUpdate.h>
            
                                // --- Project Modules ---
                                #include "config_manager.h"
                                #include "ota_handler.h"
                                #include "mqtt_logger.h"
                                // #include "bh1750_driver.h" // EXAMPLE - You must include the actual drivers you use!
            
                                // --- Client Definitions ---
                                WiFiClient localWifiClient;
                                HTTPUpdate myHttpUpdate;
                                PubSubClient localMqttClient(localWifiClient);
                                Print& logger = MqttLogger::getInstance();
            
                                // --- Callback ---
                                void local_mqtt_callback(char* topic, byte* payload, unsigned int length) {{
                                    String topicStr = String(topic);
                                    String payloadStr;
                                    payloadStr.reserve(length);
                                    for (unsigned int i = 0; i < length; i++) {{ payloadStr += (char)payload[i]; }}
                                    ota_handle_mqtt_message(topicStr, payloadStr);
                                    // Add other command handling here if needed...
                                }}
            
                                // --- Connection ---
                                void connectToLocalMqtt() {{
                                    while (!localMqttClient.connected()) {{
                                        logger.println("Attempting to connect to LOCAL MQTT Broker...");
                                        if (localMqttClient.connect(DEVICE_ID)) {{
                                            logger.println("SUCCESS: Connected to Local MQTT Broker.");
                                            // Initialize OTA after a successful connection
                                            ota_init(localWifiClient, localMqttClient, myHttpUpdate);
                                        }} else {{
                                            logger.printf("FAILED, rc=%d. Retrying in 5 seconds\\n", localMqttClient.state());
                                            delay(5000);
                                        }}
                                    }}
                                }}
            
                                // --- Main Program ---
                                void setup() {{
                                    Serial.begin(115200);
                                    delay(10);
                                    WiFi.begin(WIFI_SSID, WIFI_PASSWORD);
                                    while (WiFi.status() != WL_CONNECTED) {{ delay(500); Serial.print("."); }}
                                    Serial.println("\\nWiFi connected.");
            
                                    // Initialize logger after WiFi
                                    MqttLogger::getInstance().begin(localMqttClient, DEVICE_ID);
                                    logger.println("--- System Initializing (Local Mode) ---");
            
                                    // Configure and connect to local MQTT
                                    localMqttClient.setServer(MQTT_BROKER, MQTT_PORT);
                                    localMqttClient.setCallback(local_mqtt_callback);
                                    connectToLocalMqtt();
            
                                    // Initialize your drivers here
                                    // bh1750_setup(34); // Example driver setup
            
                                    logger.println("--- Setup Complete ---");
                                }}
            
                                void loop() {{
                                    if (!localMqttClient.connected()) {{ connectToLocalMqtt(); }}
                                    localMqttClient.loop();
                                    MqttLogger::getInstance().loop();
            
                                    // Correct non-blocking timer for periodic actions
                                    static unsigned long lastPublishTime = 0;
                                    if (millis() - lastPublishTime > 5000) {{ // 5-second interval
                                        lastPublishTime = millis();
                                        
                                        // Application logic to read sensor and publish
                                        JsonDocument doc;
                                        doc["lux"] = bh1750_read_lux(); // Example driver usage
                                        char buffer[64];
                                        serializeJson(doc, buffer);
                                        if (localMqttClient.publish("/data/local/report", buffer)) {{
                                            logger.printf("Published: %s\\n", buffer);
                                        }}
                                    }}
                                }}
                                ```
                                </CorrectExample_for_Single_Client_Architecture>
                                <OutputFormat>You MUST provide a single, complete code block for the main `.ino` file based on the single-client architecture.</OutputFormat>
                            </Prompt>
                            """)
        response = developer_model.invoke([HumanMessage(content=prompt)])
        main_code = extract_code(response.content, lang="cpp")
        completed_modules[task_id] = {"task_id": task_id, "header_code": None, "source_code": None,
                                      "main_code": main_code, "version": version}
    return {"completed_modules": completed_modules, "feedback": "", "current_module_task": task}


def integrator_node(state: AgentState) -> Dict:
    device_id = state['current_device_task']['internal_device_id']
    print(f"--- L6: INTEGRATOR: Assembling final verified firmware for '{device_id}' ---")
    project_files = state.get('project_files', {})
    final_project_files = {}
    completed_modules = state.get('completed_modules', {})
    final_project_files["lib/"] = ""
    final_project_files["src/"] = ""
    for task_id, module in completed_modules.items():
        if task_id == 'tuya_handler':
            module_dir = f"lib/{task_id}/"
            final_project_files[module_dir] = ""
            if module.get('header_code'): final_project_files[f"{module_dir}{task_id}.h"] = module['header_code']
            if module.get('source_code'):
                try:
                    source_files = json.loads(module['source_code'])
                    for filename, content in source_files.items():
                        final_project_files[f"{module_dir}{filename}"] = content
                except (json.JSONDecodeError, TypeError):
                    final_project_files[f"{module_dir}{task_id}.cpp"] = module['source_code']
            continue
        if task_id in ['config_manager', 'ota_handler', 'mqtt_logger']:
            module_dir = f"lib/{task_id}/"
            final_project_files[module_dir] = ""
            if module.get('header_code'): final_project_files[f"{module_dir}{task_id}.h"] = module['header_code']
            if module.get('source_code'): final_project_files[f"{module_dir}{task_id}.cpp"] = module['source_code']
        else:
            if module.get('header_code'): final_project_files[f"src/{task_id}.h"] = module['header_code']
            if module.get('source_code'): final_project_files[f"src/{task_id}.cpp"] = module['source_code']
            if module.get('main_code'): final_project_files[f"src/{task_id}.ino"] = module['main_code']

    user_board_model = state['current_device_task']['board']
    corrected_board_id = find_board_id(user_board_model) or "esp32dev"

    # 【性能优化】添加 lib_archive = false 来加速第三方库的编译链接过程
    final_project_files["platformio.ini"] = f"""
    [platformio]
    build_cache_dir = ../.build_cache

    [env:{device_id}]
    platform = espressif32
    board = {corrected_board_id}
    framework = arduino
    lib_deps =
        knolleary/PubSubClient
        bblanchon/ArduinoJson
    monitor_speed = 115200
    lib_extra_dirs = lib/
    lib_archive = false
    """
    project_files[device_id] = final_project_files
    return {"project_files": project_files}

def test_plan_designer_node(state: AgentState) -> Dict:
    print("--- L4: TEST PLAN DESIGNER: Creating test plan to verify MQTT JSON structure ---")
    app_main_code = state["completed_modules"].get("app_main", {}).get("main_code", "")
    current_device_role = state['current_device_task'].get('device_role', 'Unknown Device')

    prompt = textwrap.dedent(f"""
        <Prompt>
            <Role>You are a quality assurance engineer creating a precise test plan for an IoT device.</Role>
            <Goal>Analyze the provided code to find the MQTT topic and the JSON structure of the message payload. Use this to create a test plan that verifies the message format.</Goal>
            <Context>
                <DeviceRole>{current_device_role}</DeviceRole>
                <ApplicationCode>```cpp
    {app_main_code}```</ApplicationCode>
            </Context>
            <Instructions>
                1.  **Analyze `loop()`**: Focus on the code inside the `loop()` function.
                2.  **Find Publish Call**: Locate the `localMqttClient.publish(topic, payload)` call.
                3.  **Extract Topic**: Find the definition of the `topic` variable. This is the value for `device_log_topic`.
                4.  **Extract Key JSON Structure**: Examine the JSON being built for the `payload`. Identify a key part of the structure, **including quotes and colons**. This makes the check specific. For a payload `{{\\"lux\\":123}}`, the best value for `expected_log_contains` is `"lux":`.
                5.  **Construct JSON**: Fill out the test plan using the extracted topic and JSON structure.
                6.  **Handle No Publishing**: If `loop()` does not contain a recurring `localMqttClient.publish()` call, generate a test plan with an **empty `sequence` array**.
            </Instructions>
            <Example>
               <ApplicationCode>
               ...
               const char* DATA_TOPIC = "/smart_light/sensor/data";
               void loop() {{
                   ...
                   JsonDocument doc;
                   doc["illumination"] = bh1750_read_lux();
                   char buffer[64];
                   serializeJson(doc, buffer);
                   localMqttClient.publish(DATA_TOPIC, buffer);
                   ...
               }}
               </ApplicationCode>
               <Response>
               ```json
               {{
               "test_plan": {{
                   "device_log_topic": "/smart_light/sensor/data",
                   "sequence": [
                     {{
                       "name": "Check for illumination data publication",
                       "expected_log_contains": "\\"illumination\\":",
                       "timeout_seconds": 30
                     }}
                   ],
                   "success_criteria": "ALL_PASS"
               }}
               }}
               ```
               </Response>
            </Example>
            <OutputFormat>You MUST output a single, valid JSON object for the test plan.</OutputFormat>
        </Prompt>
        """)
    response = tester_model.invoke([HumanMessage(content=prompt)])
    try:
        plan_json = json.loads(extract_code(response.content, "json"))
        return {"test_plan": plan_json['test_plan']}
    except (json.JSONDecodeError, KeyError) as e:
        return {"test_plan": None, "feedback": f"FAIL: Could not generate test plan. Error: {e}"}


def deployment_and_verification_node(state: AgentState) -> Dict:
    device_task = state['current_device_task']
    device_id = device_task['internal_device_id']
    device_role = device_task.get('device_role', device_id)

    # 核心修正：创建一个对文件名安全的设备角色名
    safe_device_role = re.sub(r'[^a-zA-Z0-9_-]', '', device_role.lower().replace(' ', '_'))

    print(f"\\n--- L6: DEPLOYMENT & VERIFICATION: Preparing files for '{device_role}' ---")

    project_files = state['project_files'][device_id]

    # 核心修正：在主工作区内，为当前设备创建一个独立的子文件夹
    base_workspace_path = Path(state['workspace_path'])
    device_project_path = base_workspace_path / safe_device_role
    device_project_path.mkdir(parents=True, exist_ok=True)

    print(f"  -> Writing project files to dedicated directory: '{device_project_path}'")

    for filename, content in project_files.items():
        # 核心修正：将文件写入到设备的专属子文件夹中
        dest_path = device_project_path / filename
        if filename.endswith('/'):
            dest_path.mkdir(parents=True, exist_ok=True)
            continue
        dest_path.parent.mkdir(parents=True, exist_ok=True)
        with open(dest_path, "w", encoding="utf-8") as f:
            f.write(content)

    print(f"\n--- [VERIFICATION NODE] PHASE 3/3: Verifying device operation ---")
    test_plan = state.get('test_plan')

    if not test_plan or not isinstance(test_plan, dict) or not test_plan.get("sequence"):
        print("  -> Verification skipped (no test plan or sequence found).")
        # 核心修正：即使跳过验证，也要返回正确的 build_dir
        return {"build_dir": str(device_project_path),
                "feedback": "PASS: Verification skipped (no test plan generated)."}

    local_pc_ip = get_local_ip()
    topic_to_verify = test_plan.get("device_log_topic", f"/debug/{device_id}/log")
    print(f"  -> Test Plan found. Preparing to listen on MQTT topic: '{topic_to_verify}'")

    verifier_code = f"""
# verifier_script.py
import paho.mqtt.client as mqtt
import json, time, sys

MQTT_BROKER = "{local_pc_ip}"
MQTT_PORT = 1883
TEST_PLAN = {json.dumps(test_plan)}
DEVICE_ID = "{device_id}"
TOPIC_TO_VERIFY = "{topic_to_verify}"

test_results = {{}}
current_step_index = 0
start_time = time.time()
client = mqtt.Client(mqtt.CallbackAPIVersion.VERSION2)

def on_connect(client, userdata, flags, reason_code, properties):
    print(f"Verifier: Connected to MQTT Broker with reason code {{reason_code}}.")
    if reason_code == 0:
        print(f"Verifier: Subscribing to topic: {{TOPIC_TO_VERIFY}}")
        client.subscribe(TOPIC_TO_VERIFY)
    else:
        print("Verifier: MQTT connection failed, exiting.")
        sys.exit(1)

def on_message(client, userdata, msg):
    global current_step_index, start_time
    payload = msg.payload.decode('utf-8').strip()
    print(f"Verifier: Received message on '{{msg.topic}}': '{{payload}}'")

    if current_step_index >= len(TEST_PLAN['sequence']): return
    step = TEST_PLAN['sequence'][current_step_index]

    if step['expected_log_contains'] in payload:
        print(f"  -> MATCH FOUND for step '{{step['name']}}'!")
        test_results[step['name']] = "PASS"
        current_step_index += 1
        start_time = time.time()

client.on_connect = on_connect
client.on_message = on_message

print(f"Verifier: Connecting to {{MQTT_BROKER}}:{{MQTT_PORT}}...")
client.connect(MQTT_BROKER, MQTT_PORT, 60)
client.loop_start()

while current_step_index < len(TEST_PLAN['sequence']):
    step = TEST_PLAN['sequence'][current_step_index]
    timeout = step['timeout_seconds']
    print(f"Verifier: Waiting for message containing '{{step['expected_log_contains']}}'. Timeout in {{timeout - (time.time() - start_time):.1f}}s")
    if time.time() - start_time > timeout:
        print(f"Verifier: TIMEOUT waiting for step '{{step['name']}}'.")
        test_results[step['name']] = "FAIL: Timeout"
        break
    time.sleep(1)

client.loop_stop()
client.disconnect()
print("Verifier: Disconnected from MQTT.")

all_passed = all(res == "PASS" for res in test_results.values()) and len(test_results) == len(TEST_PLAN['sequence'])
final_result = {{"status": "PASS" if all_passed else "FAIL", "details": test_results}}

print(f"Verifier: Final Result -> {{json.dumps(final_result)}}")
with open("test_result.json", "w") as f: json.dump(final_result, f)

if not all_passed:
    sys.exit(1)
"""
    verifier_script_path = device_project_path / "run_verification.py"
    verifier_script_path.write_text(verifier_code, encoding="utf-8")

    # 核心修正：返回正确的、设备专属的文件夹路径
    return {"build_dir": str(device_project_path)}


def compile_node(state: AgentState) -> Dict:
    build_dir = Path(state["build_dir"])
    device_id = state['current_device_task']['internal_device_id']
    print(f"\n--- [COMPILE NODE] PHASE 1/3: Compiling firmware for {device_id} in '{build_dir}' ---")

    # 【诊断代码】打印关键环境变量和 pio 系统信息
    print("--- [DIAGNOSIS] Checking subprocess environment ---")
    try:
        home_dir = os.path.expanduser("~")
        print(f"  - Python's view of HOME: {home_dir}")
        print(f"  - Env Var 'HOME': {os.environ.get('HOME')}")
        print(f"  - Env Var 'USERPROFILE' (Windows): {os.environ.get('USERPROFILE')}")
        print(f"  - Env Var 'PLATFORMIO_HOME_DIR': {os.environ.get('PLATFORMIO_HOME_DIR')}")

        # 运行 pio system info 来查看 PlatformIO 是如何看待自己的环境的
        pio_info_proc = subprocess.run(
            ["platformio", "system", "info"],
            capture_output=True, text=True, encoding='utf-8', errors='ignore'
        )
        print("\n  --- pio system info output ---")
        print(pio_info_proc.stdout)
        print("  ------------------------------\n")

    except Exception as diag_e:
        print(f"  - Diagnosis step failed: {diag_e}")

    # 【诊断代码结束】
    # 性能优化：不再将所有输出捕获到内存，而是直接流式传输到控制台
    try:
        command = ["platformio", "run"]
        print(f"  -> Executing command: {' '.join(command)}")
        process = subprocess.Popen(
            command,
            cwd=build_dir,
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT, # 合并标准错误和标准输出
            text=True,
            encoding='utf-8',
            errors='ignore'
        )

        full_log = ""
        # 实时读取和打印输出
        for line in iter(process.stdout.readline, ''):
            print(line, end='') # 实时打印到控制台
            full_log += line
        process.stdout.close()
        return_code = process.wait() # 等待进程结束

        if return_code != 0:
            print(f"--- [COMPILE NODE] COMPILE FAILED with return code {return_code} ---")
            # 为修复流程返回带有 FAIL 前缀的完整日志
            return {"feedback": f"FAIL: Compile process failed.\n{full_log}"}

    except Exception as e:
        error_msg = f"FAIL: An unexpected error occurred during compilation. Exception: {type(e).__name__}: {str(e)}"
        print(f"--- [COMPILE NODE] {error_msg} ---")
        return {"feedback": error_msg}

    # 编译成功后的逻辑保持不变
    firmware_path = build_dir / ".pio" / "build" / device_id / "firmware.bin"
    if not firmware_path.exists():
        return {"feedback": f"FAIL: Compiled firmware.bin not found at {firmware_path}"}

    print(f"--- [COMPILE NODE] COMPILATION SUCCESS. Firmware ready at: {firmware_path.resolve()} ---")
    return {
        "feedback": "PASS: Compilation successful.",
        "firmware_path": str(firmware_path),
        "build_dir": str(build_dir)
    }

def usb_upload_node(state: AgentState) -> Dict:
    build_dir = Path(state["build_dir"])
    device_id = state['current_device_task']['internal_device_id']
    print(f"\n--- [USB UPLOAD NODE] PHASE 2/3: Uploading firmware via USB for {device_id} ---")

    try:
        command = ["platformio", "run", "--target", "upload"]
        print(f"  -> Executing in '{build_dir}': {' '.join(command)}")
        # 同样使用流式输出来提供实时反馈
        process = subprocess.Popen(
            command,
            cwd=build_dir,
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            text=True, encoding='utf-8', errors='ignore'
        )
        full_log = ""
        for line in iter(process.stdout.readline, ''):
            print(line, end='')
            full_log += line
        process.stdout.close()
        return_code = process.wait(timeout=300)

        if return_code != 0:
             print(f"--- [USB UPLOAD NODE] UPLOAD FAILED with return code {return_code} ---")
             return {"feedback": f"FAIL: USB Upload process failed.\n\n{full_log}"}

        print("--- [USB UPLOAD NODE] USB UPLOAD COMMAND EXECUTED ---")
        # 即使命令成功，也给予短暂延时，确保设备重启和网络连接
        print("  -> Waiting 10 seconds for device to reboot and connect to network...")
        time.sleep(10)
        return {"feedback": "PASS: USB upload command executed."}

    except subprocess.TimeoutExpired as e:
        msg = f"FAIL: USB Upload process timed out after {e.timeout} seconds."
        print(f"--- [USB UPLOAD NODE] {msg} ---")
        return {"feedback": msg}
    except Exception as e:
        msg = f"FAIL: Unexpected error during USB upload. Exception: {type(e).__name__}: {str(e)}"
        print(f"--- [USB UPLOAD NODE] {msg} ---")
        return {"feedback": msg}

def pre_deployment_pause_node(state: AgentState) -> Dict:
    print("\\n--- Waiting for user to select deployment method... ---")
    return {"available_actions": ["DEPLOY_USB", "DEPLOY_OTA"]}

def ota_deployment_node(state: AgentState) -> Dict:
    print("\\n--- Entering Node: Real OTA Deployment ---")
    build_dir = Path(state["build_dir"])
    firmware_path = Path(state["firmware_path"])
    device_id = state['current_device_task']['internal_device_id']
    local_pc_ip = get_local_ip()
    ota_pusher_code = textwrap.dedent(f"""
    import paho.mqtt.client as mqtt
    import json, time
    MQTT_BROKER = "{local_pc_ip}"
    MQTT_PORT = 1883
    DEVICE_ID = "{device_id}"
    client = mqtt.Client(mqtt.CallbackAPIVersion.VERSION2)
    client.connect(MQTT_BROKER, MQTT_PORT, 60)
    client.loop_start()
    time.sleep(1)
    if client.is_connected():
        command = {{"action": "update", "file": "firmware.bin"}}
        topic = f"/ota/{{DEVICE_ID}}/command"
        client.publish(topic, json.dumps(command))
    else:
        exit(1)
    time.sleep(1)
    client.loop_stop()
    client.disconnect()
    """)
    ota_pusher_script_path = build_dir / "ota_pusher.py"
    ota_pusher_script_path.write_text(ota_pusher_code, encoding="utf-8")
    http_server_dir = firmware_path.parent
    http_server_process = subprocess.Popen(["python", "-m", "http.server", "8000"], cwd=http_server_dir,
                                           stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
    try:
        subprocess.run(["python", ota_pusher_script_path.name], cwd=build_dir, check=True, capture_output=True,
                       text=True, encoding='utf-8', errors='ignore')
        time.sleep(20)
    except subprocess.CalledProcessError as e:
        return {"feedback": f"FAIL: OTA push script failed. Error: {e.stderr}"}
    finally:
        http_server_process.terminate()
    return {"feedback": "PASS: Real OTA deployment command sent."}


def deploy_and_verify_node(state: AgentState) -> Dict:
    print(f"\n--- [VERIFICATION NODE] PHASE 3/3: Verifying device operation ---")
    test_plan = state.get('test_plan')

    if not test_plan or not isinstance(test_plan, dict) or not test_plan.get("sequence"):
        print("  -> Verification skipped (no test plan or sequence found).")
        return {"feedback": "PASS: Verification skipped (no test plan generated)."}

    device_id = state['current_device_task']['internal_device_id']
    build_dir = Path(state["build_dir"])
    local_pc_ip = get_local_ip()

    # 核心修正：从 test_plan 动态获取要验证的 MQTT 主题
    raw_topic = test_plan.get("device_log_topic", f"/debug/{device_id}/log")
    # 确保 topic 中的 {DEVICE_ID} 占位符被正确替换
    topic_to_verify = raw_topic.format(DEVICE_ID=device_id)
    print(f"  -> Test Plan found. Preparing to listen on MQTT topic: '{topic_to_verify}'")

    verifier_code = f"""
# verifier_script.py
import paho.mqtt.client as mqtt
import json, time, sys

MQTT_BROKER = "{local_pc_ip}"
MQTT_PORT = 1883
TEST_PLAN = {json.dumps(test_plan)}
DEVICE_ID = "{device_id}"
# BUGFIX: The topic is now correctly formatted before being embedded in the script
TOPIC_TO_VERIFY = "{topic_to_verify}" 

test_results = {{}}
current_step_index = 0
start_time = time.time()
client = mqtt.Client(mqtt.CallbackAPIVersion.VERSION2)

def on_connect(client, userdata, flags, reason_code, properties):
    print(f"Verifier: Connected to MQTT Broker with reason code {{reason_code}}.")
    if reason_code == 0:
        # 核心修正：订阅正确的数据主题
        print(f"Verifier: Subscribing to topic: {{TOPIC_TO_VERIFY}}")
        client.subscribe(TOPIC_TO_VERIFY)
    else:
        print("Verifier: MQTT connection failed, exiting.")
        sys.exit(1)

def on_message(client, userdata, msg):
    global current_step_index, start_time
    payload = msg.payload.decode('utf-8').strip()
    print(f"Verifier: Received message on '{{msg.topic}}': '{{payload}}'")

    if current_step_index >= len(TEST_PLAN['sequence']): return
    step = TEST_PLAN['sequence'][current_step_index]

    if step['expected_log_contains'] in payload:
        print(f"  -> MATCH FOUND for step '{{step['name']}}'!")
        test_results[step['name']] = "PASS"
        current_step_index += 1
        start_time = time.time()

client.on_connect = on_connect
client.on_message = on_message

print(f"Verifier: Connecting to {{MQTT_BROKER}}:{{MQTT_PORT}}...")
client.connect(MQTT_BROKER, MQTT_PORT, 60)
client.loop_start()

while current_step_index < len(TEST_PLAN['sequence']):
    step = TEST_PLAN['sequence'][current_step_index]
    timeout = step['timeout_seconds']
    print(f"Verifier: Waiting for message containing '{{step['expected_log_contains']}}'. Timeout in {{timeout - (time.time() - start_time):.1f}}s")
    if time.time() - start_time > timeout:
        print(f"Verifier: TIMEOUT waiting for step '{{step['name']}}'.")
        test_results[step['name']] = "FAIL: Timeout"
        break
    time.sleep(1)

client.loop_stop()
client.disconnect()
print("Verifier: Disconnected from MQTT.")

all_passed = all(res == "PASS" for res in test_results.values()) and len(test_results) == len(TEST_PLAN['sequence'])
final_result = {{"status": "PASS" if all_passed else "FAIL", "details": test_results}}

print(f"Verifier: Final Result -> {{json.dumps(final_result)}}")
with open("test_result.json", "w") as f: json.dump(final_result, f)

if not all_passed:
    sys.exit(1)
"""
    verifier_script_path = build_dir / "run_verification.py"
    verifier_script_path.write_text(verifier_code, encoding="utf-8")

    print(f"  -> Verification script created at '{verifier_script_path}'. Executing...")
    try:
        process = subprocess.run(
            ["python", "run_verification.py"],
            cwd=build_dir,
            check=True,
            capture_output=True,
            text=True,
            encoding='utf-8',
            errors='ignore'
        )
        print("--- Verifier Script Output ---")
        print(process.stdout)
        print("----------------------------")

        with open(build_dir / "test_result.json", "r") as f:
            test_output = json.load(f)
        if test_output["status"] == "PASS":
            print("--- [VERIFICATION NODE] VERIFICATION SUCCESS ---")
            return {"feedback": "PASS: All hardware-in-the-loop tests passed."}
        else:
            details = json.dumps(test_output['details'])
            print(f"--- [VERIFICATION NODE] VERIFICATION FAILED. Details: {details} ---")
            return {"feedback": f"FAIL: Verification failed. Details: {details}"}

    except subprocess.CalledProcessError as e:
        details = f"Verification script exited with error code {e.returncode}."
        print(f"--- [VERIFICATION NODE] VERIFICATION FAILED. {details} ---")
        print("--- Verifier Script Error Output ---")
        print(e.stdout)
        print("----------------------------------")
        return {"feedback": f"FAIL: {details}\n{e.stdout}"}
# 【路由修正】check_unit_test_result 保持不变，但它的调用位置和上下文变了
def check_unit_test_result(state: AgentState) -> str:
    """决策函数：检查编译或验证结果，决定是继续还是修复。"""
    if "FAIL" in state.get('feedback', ''):
        print(f"--- [ROUTING] Feedback indicates FAILURE. Routing to REPAIR. ---")
        return "REPAIR"
    print(f"--- [ROUTING] Feedback indicates PASS. Routing to next step. ---")
    return "PASS"

# 【路由修正】deploy_and_verify_node 之后的路由
def route_after_verification(state: AgentState) -> str:
    """
    一个专门用于验证节点之后的新路由函数。
    它将区分可修复的编译失败和不可修复的验证失败。
    """
    feedback = state.get('feedback', '')
    if "FAIL: Verification failed" in feedback:
        print(
            "--- [ROUTING] Verification failed. This is a non-recoverable runtime error for the current device. Ending this device's workflow.")
        return "FINISH_DEVICE"  # 新的路由目标
    elif "FAIL" in feedback:
        # 其他类型的失败（理论上不应该在这里发生，但作为保障）
        print(f"--- [ROUTING] An unexpected failure occurred. Routing to REPAIR.")
        return "REPAIR"

    print(f"--- [ROUTING] Verification successful. Routing to DP Extractor.")
    return "PASS"


# 文件: app/langgraph_def/graph_builder.py

def dp_extractor_node(state: AgentState) -> Dict:
    """
    LLM reads the functional code of the current device and extracts Data Point (DP) information.
    [V5.2 Final Fix]: Corrects the state key to 'dp_info_list' to match the AgentState definition.
    """
    print("\\n--- Entering DP Extractor Node (V5.2 with correct state key) ---")

    current_device_task = state.get("current_device_task")
    if not current_device_task:
        return {"dp_info_list": []}

    device_id = current_device_task.get("internal_device_id")
    print(f"  -> Extracting Tuya DPs for device '{device_id}'...")

    all_project_files = state.get("project_files", {})
    actual_project_files = all_project_files.get(device_id, {})

    all_code_content = ""
    for file_name, file_content in actual_project_files.items():
        if isinstance(file_content, str):
            all_code_content += f"// --- Start of {file_name} ---\\n{file_content}\\n// --- End of {file_name} ---\\n\\n"

    if not all_code_content.strip():
        return {"dp_info_list": []}

    prompt_template = ChatPromptTemplate.from_messages([
        SystemMessagePromptTemplate.from_template(textwrap.dedent("""
            你是一个经验丰富的嵌入式系统工程师，擅长分析C/C++代码并提取设备的功能点信息。
            你的任务是根据提供的功能代码，识别其中实现的功能点（Data Point，简称DP），并严格按照指定的JSON格式输出这些功能点的信息列表。

            **！！！最重要的规则！！！**
            1.  **只看涂鸦**: 你必须只分析和涂鸦功能相关的代码。具体来说，就是寻找对 `tuya_publish_data()` 函数的调用（用于数据上报），以及在 `handle_tuya_app_commands()` 回调函数中的逻辑（用于处理指令下发）。
            2.  **必须忽略其他**: 你必须完全忽略所有与 `localMqttClient`、`PubSubClient` 的通用 `publish`/`subscribe` 方法、`logger.println` 或其他非涂鸦的通信代码。如果一个功能是通过 `localMqttClient` 或 `logger` 实现的，那它就**不是**涂鸦功能点。
            3.  **提取关键信息**: 从涂鸦相关的代码中，找到JSON负载中的`key`作为功能点的`code`(标识符)，并判断其`name`(中文名)、`type`(数据类型)和`mode`(传输方向)。

            功能点信息dp_info_list的格式是一个JSON列表，每个元素都是一个字典，代表一个功能点。
            请参考以下示例结构：
            ```json
            [
                {{
                    "id": 104,
                    "name": "亮度",
                    "code": "bright_value",
                    "mode": "rw",
                    "type": "value",
                    "define": "",
                    "remark": "",
                    "range_min": "10",
                    "range_max": "1000",
                    "step": "1",
                    "multiple": "0",
                    "unit": ""
                }}
            ]
            ```

            请严格遵循以下功能点信息规则：
            1. id：功能点ID，必填，整数，范围在101-499之间。从101开始递增生成，并避免重复。
            2. name：功能点名称，必填，根据功能代码具体实现的功能生成，支持中文。
            3. code：标识符，必填，支持英文，通常是变量名或功能函数名的小写下划线形式。
            4. mode：数据传输类型，必填。 "rw" (可上报可下发), "ro" (只上报), "wr" (只下发)。
            5. type：数据类型，必填。 "value", "string", "data", "bool", "enum"。
            6. define：数据定义。字符型填写最大长度；枚举型填写枚举值并用 "," 隔开；其他类型留空。
            7. remark：备注，默认留空。
            8. range_min/range_max/step/multiple：仅数值型必填，若无则给合理默认值。
            9. unit：单位，一般留空。

            你的输出必须是一个只包含JSON列表的字符串，不需要任何额外的解释或文本。如果代码中没有找到任何明确的**涂鸦**功能点，请返回一个空的JSON列表 `[]`。
            """)),
        HumanMessagePromptTemplate.from_template("请根据以下项目代码分析并生成功能点信息列表：\\n\\n{code_content}")
    ])

    chain = prompt_template | dp_extractor_model | JsonOutputParser()

    try:
        dp_info_list_for_device = chain.invoke({"code_content": all_code_content})
        final_list = []
        next_id = 101
        if isinstance(dp_info_list_for_device, list):
            for i, dp in enumerate(dp_info_list_for_device):
                if isinstance(dp, dict) and "code" in dp and "name" in dp:
                    dp['id'] = next_id + i
                    for key in ["mode", "type", "define", "remark", "range_min", "range_max", "step", "multiple", "unit"]:
                        if key not in dp:
                            dp[key] = ""
                    final_list.append(dp)

        print(f"  -> Successfully extracted and validated {len(final_list)} Tuya DP(s):")
        print(json.dumps(final_list, indent=2, ensure_ascii=False))
        # 核心修正：使用 AgentState 中定义的 "dp_info_list" 作为键名
        return {"dp_info_list": final_list}
    except Exception as e:
        print(f"  -> ERROR: Failed to extract or parse DPs: {e}")
        # 核心修正：即使失败也要使用正确的键名返回空列表
        return {"dp_info_list": []}


def device_artifact_generator_node(state: AgentState) -> Dict:
    """
    在单个设备处理流程结束后，为其生成专属的产出物（如 .xlsx 文件）。
    [V1.1 修正版]: 修正了读取状态的键名。
    """
    print("\\n--- [DEVICE ARTIFACT GENERATOR]: Generating artifacts for current device ---")

    workspace_path = Path(state['workspace_path'])
    current_device = state.get('current_device_task', {})
    device_role = current_device.get('device_role', 'unknown_device')

    # 核心修正：从 AgentState 中定义的 "dp_info_list" 读取数据
    dp_list = state.get('dp_info_list', [])

    if not dp_list:
        print(f"  -> No data points extracted for device '{device_role}'. Skipping Excel file generation.")
        return {}

    safe_device_role = "".join(c for c in device_role if c.isalnum() or c in (' ', '_')).rstrip()
    filename = workspace_path / f"{safe_device_role}_dps.xlsx"
    print(f"  -> Generating Tuya DP template for '{device_role}' at: {filename}")

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "功能点 (DP)"

        headers = [
            "DP ID", "功能点名称", "标识符", "数据传输类型", "数据类型",
            "数据定义", "备注", "数据范围-最小值", "数据范围-最大值",
            "间距", "倍数", "单位"
        ]
        ws.append(headers)

        column_widths = [10, 20, 20, 15, 12, 30, 20, 15, 15, 10, 10, 10]
        for i, column_width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = column_width

        for dp in dp_list:
            row_data = [
                dp.get("id"), dp.get("name"), dp.get("code"),
                dp.get("mode"), dp.get("type"), dp.get("define"),
                dp.get("remark"), dp.get("range_min"), dp.get("range_max"),
                dp.get("step"), dp.get("multiple"), dp.get("unit")
            ]
            ws.append(row_data)

        wb.save(filename)
        print(f"  -> Successfully saved Excel file for '{device_role}' with {len(dp_list)} data points.")

    except Exception as e:
        print(f"  -> ERROR: Failed to generate Excel file for '{device_role}': {e}")

    # 核心修正：清空状态时也使用正确的键名
    return {"dp_info_list": []}

# =================================================================================
# 5. Graph Definition & Logic
# =================================================================================

def check_device_queue(state: AgentState) -> str:
    """【修正】决策函数：检查是否还有待处理的设备任务。"""
    # 在开始处理一个新设备前，清理上个设备可能留下的旧状态
    state['feedback'] = ""
    state['user_action'] = None
    state['deployment_choice'] = None
    if state.get("current_device_task"):
        return "continue_to_development"
    return "finish_all_devices"

def check_module_queue(state: AgentState) -> str:
    """决策函数：检查是否还有待处理的模块任务。"""
    if state.get("current_module_task"):
        return "continue_development"
    return "finish_development"

def check_unit_test_result(state: AgentState) -> str:
    """决策函数：检查编译或验证结果，决定是继续还是修复。"""
    if "FAIL" in state.get('feedback', ''):
        print(f"--- [ROUTING] Feedback indicates FAILURE. Routing to REPAIR. ---")
        return "REPAIR"
    print(f"--- [ROUTING] Feedback indicates PASS. Routing to next step. ---")
    return "PASS"

def master_router_node(state: AgentState) -> dict:
    """【修正】图的总入口节点，仅作为正式的节点存在，返回空字典。"""
    print("--- [MASTER ROUTER] Evaluating entry point... ---")
    return {}

def master_router_logic(state: AgentState) -> str:
    """【修正】用于总入口节点的路由决策函数。"""
    if state.get("user_action"):
        print("--- [ROUTING LOGIC] User action found. Resuming from pause. ---")
        return "resume_from_pause"
    else:
        print("--- [ROUTING LOGIC] No user action. Starting from beginning. ---")
        return "start_from_beginning"

# 【新增】一个合格的、用于恢复流程的 “工作” 节点
def resume_router_node(state: AgentState) -> dict:
    """
    这是一个合格的图节点。它的工作很简单，就是打印一条日志。
    它将在 master_router 决定恢复流程后被调用。
    """
    print("--- [RESUME NODE] Workflow is resuming. Preparing to route deployment action... ---")
    return {}  # 作为一个合格的节点，它返回一个字典


# 【新增】一个合格的、用于恢复流程的 “决策” 函数
def route_deployment_logic(state: AgentState) -> str:
    """
    这是一个合格的条件边函数。它检查用户的操作并返回一个字符串决策。
    """
    user_action = state.get("user_action")
    if user_action == 'DEPLOY_USB':
        print("--- [ROUTING LOGIC] User chose USB. Routing to usb_upload_node. ---")
        return "REAL_USB_DEPLOY"
    elif user_action == 'DEPLOY_OTA':
        print("--- [ROUTING LOGIC] User chose OTA. Routing to ota_deployment_node. ---")
        return "REAL_OTA_DEPLOY"

    # 异常情况，理论上不应发生，但作为保护
    print("--- [ROUTING LOGIC] No user action found in a resumed state. Ending deployment phase. ---")
    return "END_DEPLOYMENT"

def build_graph():
    """
    【V3.2 架构修正】构建支持“暂停-恢复”生命周期的工作流图。
    """
    workflow = StateGraph(AgentState)

    # 添加所有节点
    workflow.add_node("master_router", master_router_node)
    workflow.add_node("plan_enrichment_node", plan_enrichment_node)
    workflow.add_node("device_dispatcher", device_dispatcher_node)
    workflow.add_node("module_architect", module_architect_node)
    workflow.add_node("module_dispatcher", module_dispatcher_node)
    workflow.add_node("api_designer", api_designer_node)
    workflow.add_node("developer", developer_node)
    workflow.add_node("integrator", integrator_node)
    workflow.add_node("test_plan_designer", test_plan_designer_node)
    workflow.add_node("deployment_and_verification", deployment_and_verification_node)
    workflow.add_node("compile_node", compile_node)
    workflow.add_node("pre_deployment_pause", pre_deployment_pause_node)
    workflow.add_node("usb_upload_node", usb_upload_node)
    workflow.add_node("ota_deployment_node", ota_deployment_node)
    workflow.add_node("deploy_and_verify_node", deploy_and_verify_node)
    workflow.add_node("dp_extractor", dp_extractor_node)
    workflow.add_node("device_artifact_generator", device_artifact_generator_node) # <-- 添加新节点

    # 【新增】添加我们新的、用于恢复的节点
    workflow.add_node("resume_router", resume_router_node)

    # 设定图的唯一入口
    workflow.set_entry_point("master_router")

    # --- 核心路由逻辑 ---

    # 1. 从总入口 master_router 开始决策
    workflow.add_conditional_edges(
        "master_router",
        master_router_logic,
        {
            # 【核心修改】如果是新开始，先走需求细化，再走设备分发
            "start_from_beginning": "plan_enrichment_node",
            "resume_from_pause": "resume_router"
        }
    )

    workflow.add_edge("plan_enrichment_node", "device_dispatcher")

    # 2. 从 resume_router 节点出发，进行部署方式的决策
    workflow.add_conditional_edges(
        "resume_router",
        route_deployment_logic,  # 使用我们新的、合格的决策函数
        {
            "REAL_USB_DEPLOY": "usb_upload_node",
            "REAL_OTA_DEPLOY": "ota_deployment_node",
            "END_DEPLOYMENT": END  # 异常情况则结束
        }
    )

    # --- 其余的图结构保持不变 ---

    # 正常的设备处理循环
    # [最终修正] 调整设备处理循环和结束逻辑
    workflow.add_conditional_edges(
        "device_dispatcher",
        check_device_queue,
        {
            "continue_to_development": "module_architect",
            "finish_all_devices": END  # <-- 当设备队列为空时，直接结束工作流
        }
    )

    # 模块开发循环
    workflow.add_edge("module_architect", "module_dispatcher")
    workflow.add_conditional_edges(
        "module_dispatcher",
        check_module_queue,
        {"continue_development": "api_designer", "finish_development": "integrator"}
    )
    workflow.add_edge("api_designer", "developer")
    workflow.add_edge("developer", "module_dispatcher")

    # 集成、测试、编译
    workflow.add_edge("integrator", "test_plan_designer")
    workflow.add_edge("test_plan_designer", "deployment_and_verification")
    workflow.add_edge("deployment_and_verification", "compile_node")

    # 编译后的路由
    workflow.add_conditional_edges(
        "compile_node",
        check_unit_test_result,
        {
            "PASS": "pre_deployment_pause",
            "REPAIR": "developer"
        }
    )

    # 暂停节点是此阶段的终点
    workflow.add_edge("pre_deployment_pause", END)

    # 部署后的流程
    workflow.add_edge("usb_upload_node", "deploy_and_verify_node")
    workflow.add_edge("ota_deployment_node", "deploy_and_verify_node")

    # --- 【架构核心修正】验证后的路由逻辑 ---
    workflow.add_conditional_edges(
        "deploy_and_verify_node",
        route_after_verification,  # 使用新的、更智能的路由函数
        {
            "PASS": "dp_extractor",  # 验证成功，继续
            "REPAIR": "developer",  # 其他失败，尝试修复
            "FINISH_DEVICE": "device_dispatcher"  # 验证失败，处理下一个设备
        }
    )

    # [最终修正] 最后的流程：提取DP -> 生成文件 -> 返回分发器
    workflow.add_edge("dp_extractor", "device_artifact_generator")
    workflow.add_edge("device_artifact_generator", "device_dispatcher")

    # 编译并返回图
    compiled_graph = workflow.compile()
    return compiled_graph